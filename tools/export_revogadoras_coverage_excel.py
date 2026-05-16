import argparse
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config import get_chunks_path, get_documents_dir, get_registry_path


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)

MONTHS_PT = {
    "janeiro": 1,
    "fevereiro": 2,
    "marco": 3,
    "marcoo": 3,
    "abril": 4,
    "maio": 5,
    "junho": 6,
    "julho": 7,
    "agosto": 8,
    "setembro": 9,
    "outubro": 10,
    "novembro": 11,
    "dezembro": 12,
}

TYPE_PATTERNS: List[Tuple[str, str]] = [
    ("emenda constitucional", "Emenda Constitucional"),
    ("instrucao normativa", "Instrucao Normativa"),
    ("ordem de servico", "Ordem de Servico"),
    ("medida provisoria", "Medida Provisoria"),
    ("decreto-lei", "Decreto-Lei"),
    ("decreto lei", "Decreto-Lei"),
    ("decreto", "Decreto"),
    ("resolucao anm", "Resolucao"),
    ("resolucao", "Resolucao"),
    ("portaria", "Portaria"),
    ("lei complementar", "Lei Complementar"),
    ("lei ordinaria", "Lei"),
    ("lei", "Lei"),
    ("deliberacao", "Deliberacao"),
    ("circular", "Circular"),
    ("alvara", "Alvara"),
]


def _maybe_fix_mojibake(text: Any) -> str:
    if text is None:
        return ""
    value = str(text)
    if not any(marker in value for marker in ("Ã", "Â", "â", "ï¿½")):
        return value
    try:
        return value.encode("latin-1").decode("utf-8")
    except Exception:
        return value


def _normalize_spaces(text: Any) -> str:
    return re.sub(r"\s+", " ", _maybe_fix_mojibake(text)).strip()


def _strip_accents(text: Any) -> str:
    normalized = unicodedata.normalize("NFKD", _normalize_spaces(text))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_lookup_text(text: Any) -> str:
    value = _strip_accents(text).lower()
    value = value.replace("\\", "/")
    value = value.split("/")[-1]
    if value.endswith(".pdf"):
        value = value[:-4]
    value = re.sub(r"[_-](\d+)$", "", value)
    value = re.sub(r"[^a-z0-9]+", "", value)
    return value


def _normalize_tipo_norma(text: Any) -> str:
    value = _strip_accents(text).lower()
    value = _normalize_spaces(value)
    value = re.sub(r"\banm\b", "", value)
    value = _normalize_spaces(value)
    for raw, canonical in TYPE_PATTERNS:
        if raw == value:
            return canonical
    return value.title() if value else ""


def _normalize_numero_norma(value: Any) -> str:
    text = _strip_accents(value).lower()
    text = text.replace(".", "")
    text = re.sub(r"[^0-9a-z/-]+", "", text)
    if "/" in text:
        text = text.split("/")[0]
    if "-" in text:
        text = text.split("-")[0]
    return text.strip()


def _parse_year(value: Any) -> Optional[int]:
    match = re.search(r"\b(19\d{2}|20\d{2})\b", _strip_accents(value))
    if not match:
        return None
    return int(match.group(1))


def _parse_pt_date_to_iso(value: Any) -> Optional[str]:
    text = _strip_accents(value).lower()
    if not text:
        return None

    match = re.search(r"\b(\d{1,2})[./-](\d{1,2})[./-]((?:19|20)\d{2}|\d{2})\b", text)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year = int(match.group(3))
        year = year if year >= 100 else (1900 + year if year >= 50 else 2000 + year)
        try:
            return datetime(year, month, day).date().isoformat()
        except ValueError:
            return None

    match = re.search(
        r"\b(\d{1,2})\s+de\s+([a-zç]+)\s+(?:de\s+)?((?:19|20)\d{2})\b",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        return None

    day = int(match.group(1))
    month_name = match.group(2).replace("ç", "c")
    month = MONTHS_PT.get(month_name)
    year = int(match.group(3))
    if not month:
        return None
    try:
        return datetime(year, month, day).date().isoformat()
    except ValueError:
        return None


def _detect_tipo_norma(text: Any) -> str:
    normalized = _strip_accents(text).lower()
    for raw, canonical in TYPE_PATTERNS:
        if re.search(rf"\b{re.escape(raw)}\b", normalized):
            return canonical
    return ""


def _extract_norm_reference(text: Any) -> Dict[str, Any]:
    raw = _normalize_spaces(text)
    normalized = _strip_accents(raw).lower()

    tipo = _detect_tipo_norma(raw)
    numero = None
    ano = None

    match_num_year = re.search(r"\b(\d{1,6}(?:\.\d{3})?)\s*[/_-]\s*(19\d{2}|20\d{2})\b", normalized)
    if match_num_year:
        numero = match_num_year.group(1)
        ano = int(match_num_year.group(2))
    else:
        match_after_tipo = None
        if tipo:
            tipo_norm = _strip_accents(tipo).lower().replace("-", r"[- ]")
            match_after_tipo = re.search(
                rf"\b{tipo_norm}\b[^0-9]{{0,20}}(?:n[º°o.]{{0,3}}\s*)?(\d{{1,6}}(?:\.\d{{3}})?)\b",
                normalized,
            )
        if match_after_tipo:
            numero = match_after_tipo.group(1)
        else:
            match_num = re.search(r"\bn[º°o.]?\s*(\d{1,6}(?:\.\d{3})?)\b", normalized)
            if match_num:
                numero = match_num.group(1)

        ano = _parse_year(normalized)

    date_iso = _parse_pt_date_to_iso(raw)
    if date_iso and not ano:
        ano = int(date_iso[:4])

    numero_norm = _normalize_numero_norma(numero)
    tipo_norm = _normalize_tipo_norma(tipo)
    key_full = f"{tipo_norm}|{numero_norm}|{ano}" if tipo_norm and numero_norm and ano else ""

    return {
        "raw": raw,
        "tipo_norma": tipo_norm,
        "numero_norma": numero_norm,
        "ano_norma": ano,
        "data_norma": date_iso or "",
        "tem_chave_completa": bool(key_full),
        "chave_completa": key_full,
    }


def _select_best_candidate(current: Optional[Dict[str, Any]], candidate: Dict[str, Any]) -> Dict[str, Any]:
    if current is None:
        return candidate

    mode_rank = {"line": 3, "": 2, None: 2, "inline": 1}

    def score(row: Dict[str, Any]) -> Tuple[int, int, int, int, int]:
        parsed = row["parsed_revogadora"]
        return (
            mode_rank.get(row.get("parse_mode"), 0),
            1 if parsed.get("tem_chave_completa") else 0,
            1 if parsed.get("data_norma") or row.get("revogado_por_data") else 0,
            1 if row.get("confianca_parse_revogacao") not in (None, "") else 0,
            -len(parsed.get("raw") or ""),
        )

    return candidate if score(candidate) > score(current) else current


def _iter_chunks(path: Path) -> Iterable[Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        for line_no, raw in enumerate(handle, start=1):
            line = raw.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSON invalido em {path} na linha {line_no}: {exc}") from exc


def _load_registry_entries(path: Path) -> List[str]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        return [_maybe_fix_mojibake(name) for name in data.keys()]
    return []


def _load_physical_docs(path: Path) -> List[str]:
    if not path.exists():
        return []
    return [p.name for p in sorted(path.rglob("*.pdf"))]


def _build_doc_indexes(names: Iterable[str]) -> Dict[str, Any]:
    by_full_key: Dict[str, List[str]] = defaultdict(list)
    by_type_num: Dict[str, List[str]] = defaultdict(list)
    normalized_names: Dict[str, List[str]] = defaultdict(list)

    for name in names:
        fixed_name = _maybe_fix_mojibake(name)
        normalized_names[_normalize_lookup_text(fixed_name)].append(fixed_name)
        parsed = _extract_norm_reference(fixed_name)
        full_key = parsed.get("chave_completa")
        if full_key:
            by_full_key[full_key].append(fixed_name)
        type_num = ""
        if parsed.get("tipo_norma") and parsed.get("numero_norma"):
            type_num = f"{parsed['tipo_norma']}|{parsed['numero_norma']}"
            by_type_num[type_num].append(fixed_name)

    return {
        "by_full_key": by_full_key,
        "by_type_num": by_type_num,
        "normalized_names": normalized_names,
    }


def _match_presence(parsed_ref: Dict[str, Any], indexes: Dict[str, Any]) -> Tuple[bool, str, str]:
    full_key = parsed_ref.get("chave_completa") or ""
    if full_key and full_key in indexes["by_full_key"]:
        matches = sorted(set(indexes["by_full_key"][full_key]))
        return True, "chave_completa", " | ".join(matches[:5])

    type_num = ""
    if parsed_ref.get("tipo_norma") and parsed_ref.get("numero_norma"):
        type_num = f"{parsed_ref['tipo_norma']}|{parsed_ref['numero_norma']}"
    if type_num and type_num in indexes["by_type_num"]:
        matches = sorted(set(indexes["by_type_num"][type_num]))
        unique_norms = {
            _extract_norm_reference(item).get("chave_completa") or item
            for item in matches
        }
        if len(unique_norms) == 1:
            return True, "tipo_numero_unico", " | ".join(matches[:5])

    normalized_raw = _normalize_lookup_text(parsed_ref.get("raw"))
    if normalized_raw and normalized_raw in indexes["normalized_names"]:
        matches = sorted(set(indexes["normalized_names"][normalized_raw]))
        return True, "nome_normalizado", " | ".join(matches[:5])

    return False, "", ""


def _quality_label(row: Dict[str, Any]) -> str:
    parsed = row["parsed_revogadora"]
    mode = row.get("parse_mode")
    if mode == "line" and parsed.get("tem_chave_completa"):
        return "alta"
    if parsed.get("tem_chave_completa") and mode != "inline":
        return "media"
    return "baixa"


def _load_best_revocation_rows(chunks_path: Path) -> List[Dict[str, Any]]:
    best_by_doc: Dict[str, Dict[str, Any]] = {}
    revoked_docs_counter = Counter()

    for chunk in _iter_chunks(chunks_path):
        doc_name = _maybe_fix_mojibake(chunk.get("doc_name") or chunk.get("doc") or "")
        revogado_por = _normalize_spaces(chunk.get("revogado_por"))
        if not doc_name or not revogado_por:
            continue

        revoked_docs_counter[doc_name] += 1
        parsed_revogadora = _extract_norm_reference(revogado_por)
        source_parsed = _extract_norm_reference(doc_name)

        candidate = {
            "documento_revogado": doc_name,
            "documento_revogado_doc_id": chunk.get("doc_id") or "",
            "documento_revogado_status": _normalize_spaces(chunk.get("status_normativo")) or "",
            "documento_revogado_tipo_revogacao": _normalize_spaces(chunk.get("tipo_revogacao")) or "",
            "documento_revogado_data_inicio_vigencia": _normalize_spaces(chunk.get("data_inicio_vigencia")) or "",
            "documento_revogado_data_fim_vigencia": _normalize_spaces(chunk.get("data_fim_vigencia")) or "",
            "documento_revogado_tipo_norma": source_parsed.get("tipo_norma") or "",
            "documento_revogado_numero_norma": source_parsed.get("numero_norma") or "",
            "documento_revogado_ano_norma": source_parsed.get("ano_norma") or "",
            "revogadora_raw": parsed_revogadora.get("raw") or "",
            "revogado_por_data": _normalize_spaces(chunk.get("revogado_por_data")) or "",
            "revogado_por_doc_id_parser": _normalize_spaces(chunk.get("revogado_por_doc_id")) or "",
            "parse_mode": chunk.get("parse_revogacao_match_mode") or "",
            "confianca_parse_revogacao": chunk.get("confianca_parse_revogacao"),
            "parsed_revogadora": parsed_revogadora,
        }

        best_by_doc[doc_name] = _select_best_candidate(best_by_doc.get(doc_name), candidate)

    rows: List[Dict[str, Any]] = []
    for row in best_by_doc.values():
        row["documento_revogado_qtd_chunks_com_revogado_por"] = revoked_docs_counter[row["documento_revogado"]]
        row["qualidade_sinal"] = _quality_label(row)
        rows.append(row)

    rows.sort(key=lambda item: (item["documento_revogado"]))
    return rows


def _build_detail_rows(
    revocation_rows: List[Dict[str, Any]],
    registry_index: Dict[str, Any],
    physical_index: Dict[str, Any],
    chunks_index: Dict[str, Any],
) -> List[Dict[str, Any]]:
    detail_rows: List[Dict[str, Any]] = []

    for row in revocation_rows:
        parsed = row["parsed_revogadora"]
        in_registry, strategy_registry, matches_registry = _match_presence(parsed, registry_index)
        in_physical, strategy_physical, matches_physical = _match_presence(parsed, physical_index)
        in_chunks, strategy_chunks, matches_chunks = _match_presence(parsed, chunks_index)

        detail_rows.append(
            {
                "documento_revogado": row["documento_revogado"],
                "tipo_norma_revogada": row["documento_revogado_tipo_norma"],
                "numero_norma_revogada": row["documento_revogado_numero_norma"],
                "ano_norma_revogada": row["documento_revogado_ano_norma"],
                "status_normativo": row["documento_revogado_status"],
                "tipo_revogacao": row["documento_revogado_tipo_revogacao"],
                "data_inicio_vigencia": row["documento_revogado_data_inicio_vigencia"],
                "data_fim_vigencia": row["documento_revogado_data_fim_vigencia"],
                "revogadora_raw": row["revogadora_raw"],
                "tipo_norma_revogadora": parsed.get("tipo_norma") or "",
                "numero_norma_revogadora": parsed.get("numero_norma") or "",
                "ano_norma_revogadora": parsed.get("ano_norma") or "",
                "data_revogadora": parsed.get("data_norma") or row["revogado_por_data"],
                "chave_revogadora": parsed.get("chave_completa") or "",
                "tem_chave_completa": "sim" if parsed.get("tem_chave_completa") else "nao",
                "parse_mode": row["parse_mode"] or "",
                "qualidade_sinal": row["qualidade_sinal"],
                "confianca_parse_revogacao": row["confianca_parse_revogacao"],
                "revogado_por_doc_id_parser": row["revogado_por_doc_id_parser"],
                "chunks_com_mesma_revogadora_no_doc": row["documento_revogado_qtd_chunks_com_revogado_por"],
                "presente_no_chunks": "sim" if in_chunks else "nao",
                "match_chunks": strategy_chunks,
                "documentos_chunks_encontrados": matches_chunks,
                "presente_no_registry": "sim" if in_registry else "nao",
                "match_registry": strategy_registry,
                "documentos_registry_encontrados": matches_registry,
                "presente_na_pasta_documentos": "sim" if in_physical else "nao",
                "match_pasta_documentos": strategy_physical,
                "documentos_pasta_encontrados": matches_physical,
            }
        )

    detail_rows.sort(
        key=lambda item: (
            item["presente_na_pasta_documentos"],
            item["presente_no_chunks"],
            item["presente_no_registry"],
            item["qualidade_sinal"],
            item["documento_revogado"],
        )
    )
    return detail_rows


def _group_distinct_revogadoras(detail_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}

    for row in detail_rows:
        key = row["chave_revogadora"] or f"raw::{_normalize_lookup_text(row['revogadora_raw'])}"
        group = grouped.get(key)
        if group is None:
            group = {
                "revogadora_raw_representativa": row["revogadora_raw"],
                "tipo_norma_revogadora": row["tipo_norma_revogadora"],
                "numero_norma_revogadora": row["numero_norma_revogadora"],
                "ano_norma_revogadora": row["ano_norma_revogadora"],
                "data_revogadora": row["data_revogadora"],
                "chave_revogadora": row["chave_revogadora"],
                "tem_chave_completa": row["tem_chave_completa"],
                "qualidade_maxima": row["qualidade_sinal"],
                "parse_modes": set(),
                "documentos_revogados": [],
                "present_chunks": row["presente_no_chunks"] == "sim",
                "present_registry": row["presente_no_registry"] == "sim",
                "present_physical": row["presente_na_pasta_documentos"] == "sim",
                "matches_chunks": set(),
                "matches_registry": set(),
                "matches_physical": set(),
            }
            grouped[key] = group

        group["parse_modes"].add(row["parse_mode"] or "")
        group["documentos_revogados"].append(row["documento_revogado"])
        group["present_chunks"] = group["present_chunks"] or row["presente_no_chunks"] == "sim"
        group["present_registry"] = group["present_registry"] or row["presente_no_registry"] == "sim"
        group["present_physical"] = group["present_physical"] or row["presente_na_pasta_documentos"] == "sim"

        if row["documentos_chunks_encontrados"]:
            group["matches_chunks"].update(row["documentos_chunks_encontrados"].split(" | "))
        if row["documentos_registry_encontrados"]:
            group["matches_registry"].update(row["documentos_registry_encontrados"].split(" | "))
        if row["documentos_pasta_encontrados"]:
            group["matches_physical"].update(row["documentos_pasta_encontrados"].split(" | "))

        quality_rank = {"baixa": 1, "media": 2, "alta": 3}
        if quality_rank.get(row["qualidade_sinal"], 0) > quality_rank.get(group["qualidade_maxima"], 0):
            group["qualidade_maxima"] = row["qualidade_sinal"]
        if len(row["revogadora_raw"]) < len(group["revogadora_raw_representativa"]):
            group["revogadora_raw_representativa"] = row["revogadora_raw"]

    distinct_rows: List[Dict[str, Any]] = []
    for group in grouped.values():
        distinct_rows.append(
            {
                "revogadora_raw_representativa": group["revogadora_raw_representativa"],
                "tipo_norma_revogadora": group["tipo_norma_revogadora"],
                "numero_norma_revogadora": group["numero_norma_revogadora"],
                "ano_norma_revogadora": group["ano_norma_revogadora"],
                "data_revogadora": group["data_revogadora"],
                "chave_revogadora": group["chave_revogadora"],
                "tem_chave_completa": group["tem_chave_completa"],
                "qualidade_maxima": group["qualidade_maxima"],
                "parse_modes": ", ".join(sorted(mode for mode in group["parse_modes"] if mode)) or "(vazio)",
                "qtd_normas_revogadas": len(group["documentos_revogados"]),
                "documentos_revogados": " | ".join(sorted(group["documentos_revogados"])),
                "presente_no_chunks": "sim" if group["present_chunks"] else "nao",
                "documentos_chunks_encontrados": " | ".join(sorted(group["matches_chunks"])[:8]),
                "presente_no_registry": "sim" if group["present_registry"] else "nao",
                "documentos_registry_encontrados": " | ".join(sorted(group["matches_registry"])[:8]),
                "presente_na_pasta_documentos": "sim" if group["present_physical"] else "nao",
                "documentos_pasta_encontrados": " | ".join(sorted(group["matches_physical"])[:8]),
            }
        )

    distinct_rows.sort(
        key=lambda item: (
            item["presente_na_pasta_documentos"],
            item["presente_no_chunks"],
            item["presente_no_registry"],
            item["qualidade_maxima"],
            -(item["qtd_normas_revogadas"]),
            item["revogadora_raw_representativa"],
        )
    )
    return distinct_rows


def _write_sheet(ws, rows: List[Dict[str, Any]], title: str) -> None:
    ws.title = title
    if not rows:
        ws.append(["sem_dados"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths: Dict[int, int] = {}
    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows[:500]:
            max_len = max(max_len, len(str(row.get(header, ""))))
        widths[idx] = min(max(max_len + 2, 14), 60)

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = BODY_ALIGNMENT


def _build_summary_rows(
    base_name: str,
    chunks_path: Path,
    registry_path: Path,
    docs_dir: Path,
    chunks_doc_names: List[str],
    registry_names: List[str],
    physical_names: List[str],
    detail_rows: List[Dict[str, Any]],
    distinct_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    return [
        {"metrica": "base_analisada", "valor": base_name},
        {"metrica": "chunks_path", "valor": str(chunks_path)},
        {"metrica": "registry_path", "valor": str(registry_path)},
        {"metrica": "documentos_dir", "valor": str(docs_dir)},
        {"metrica": "documentos_indexados_no_chunks", "valor": len(chunks_doc_names)},
        {"metrica": "documentos_no_registry", "valor": len(registry_names)},
        {"metrica": "pdfs_na_pasta_documentos", "valor": len(physical_names)},
        {"metrica": "documentos_revogados_com_revogado_por", "valor": len(detail_rows)},
        {"metrica": "normas_revogadoras_distintas", "valor": len(distinct_rows)},
        {
            "metrica": "revogadoras_distintas_presentes_no_chunks",
            "valor": sum(1 for row in distinct_rows if row["presente_no_chunks"] == "sim"),
        },
        {
            "metrica": "revogadoras_distintas_ausentes_no_chunks",
            "valor": sum(1 for row in distinct_rows if row["presente_no_chunks"] != "sim"),
        },
        {
            "metrica": "revogadoras_distintas_presentes_no_registry",
            "valor": sum(1 for row in distinct_rows if row["presente_no_registry"] == "sim"),
        },
        {
            "metrica": "revogadoras_distintas_ausentes_no_registry",
            "valor": sum(1 for row in distinct_rows if row["presente_no_registry"] != "sim"),
        },
        {
            "metrica": "revogadoras_distintas_presentes_na_pasta_documentos",
            "valor": sum(1 for row in distinct_rows if row["presente_na_pasta_documentos"] == "sim"),
        },
        {
            "metrica": "revogadoras_distintas_ausentes_na_pasta_documentos",
            "valor": sum(1 for row in distinct_rows if row["presente_na_pasta_documentos"] != "sim"),
        },
        {
            "metrica": "revogadoras_distintas_com_chave_completa",
            "valor": sum(1 for row in distinct_rows if row["tem_chave_completa"] == "sim"),
        },
        {
            "metrica": "revogadoras_distintas_baixa_qualidade",
            "valor": sum(1 for row in distinct_rows if row["qualidade_maxima"] == "baixa"),
        },
        {
            "metrica": "observacao",
            "valor": (
                "A pasta fisica de documentos foi comparada separadamente da base indexada. "
                "Se houver divergencia grande entre pasta e chunks/registry, isso indica base indexada "
                "mais ampla do que a pasta atualmente disponivel."
            ),
        },
    ]


def export_revogadoras_coverage(base_name: str, output_path: Path, docs_dir_override: Optional[Path] = None) -> Dict[str, Any]:
    chunks_path = get_chunks_path(base_name)
    registry_path = get_registry_path(base_name)
    docs_dir = docs_dir_override or get_documents_dir(base_name)

    if not chunks_path.exists():
        raise FileNotFoundError(f"chunks.jsonl nao encontrado: {chunks_path}")
    if not registry_path.exists():
        raise FileNotFoundError(f"registry.json nao encontrado: {registry_path}")

    revocation_rows = _load_best_revocation_rows(chunks_path)
    chunks_doc_names = sorted({row["documento_revogado"] for row in revocation_rows} | {
        _maybe_fix_mojibake(chunk.get("doc_name") or chunk.get("doc") or "")
        for chunk in _iter_chunks(chunks_path)
        if _maybe_fix_mojibake(chunk.get("doc_name") or chunk.get("doc") or "")
    })
    registry_names = sorted(set(_load_registry_entries(registry_path)))
    physical_names = sorted(set(_load_physical_docs(docs_dir)))

    chunks_index = _build_doc_indexes(chunks_doc_names)
    registry_index = _build_doc_indexes(registry_names)
    physical_index = _build_doc_indexes(physical_names)

    detail_rows = _build_detail_rows(revocation_rows, registry_index, physical_index, chunks_index)
    distinct_rows = _group_distinct_revogadoras(detail_rows)

    missing_chunks = [row for row in distinct_rows if row["presente_no_chunks"] != "sim"]
    missing_registry = [row for row in distinct_rows if row["presente_no_registry"] != "sim"]
    missing_physical = [row for row in distinct_rows if row["presente_na_pasta_documentos"] != "sim"]
    low_quality = [row for row in distinct_rows if row["qualidade_maxima"] == "baixa"]

    summary_rows = _build_summary_rows(
        base_name=base_name,
        chunks_path=chunks_path,
        registry_path=registry_path,
        docs_dir=docs_dir,
        chunks_doc_names=chunks_doc_names,
        registry_names=registry_names,
        physical_names=physical_names,
        detail_rows=detail_rows,
        distinct_rows=distinct_rows,
    )

    wb = Workbook()
    _write_sheet(wb.active, summary_rows, "resumo")
    _write_sheet(wb.create_sheet(), distinct_rows, "revogadoras_distintas")
    _write_sheet(wb.create_sheet(), missing_physical, "ausentes_pasta_docs")
    _write_sheet(wb.create_sheet(), missing_chunks, "ausentes_chunks")
    _write_sheet(wb.create_sheet(), missing_registry, "ausentes_registry")
    _write_sheet(wb.create_sheet(), low_quality, "baixa_qualidade")
    _write_sheet(wb.create_sheet(), detail_rows, "detalhe_revogados")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "output_path": output_path,
        "summary_rows": summary_rows,
        "detail_rows": detail_rows,
        "distinct_rows": distinct_rows,
        "missing_chunks": missing_chunks,
        "missing_registry": missing_registry,
        "missing_physical": missing_physical,
        "low_quality": low_quality,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Lista normas revogadoras extraidas do parser de revogacao e cruza "
            "com a base indexada e a pasta fisica de documentos."
        )
    )
    parser.add_argument("--base", required=True, help="Nome da base (ex.: ANM_Legis_tratada2).")
    parser.add_argument("--docs-dir", type=Path, help="Pasta fisica de documentos para override do caminho padrao.")
    parser.add_argument("--output", type=Path, help="Arquivo .xlsx de saida.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = args.output or (PROJECT_ROOT / "reports" / f"revogadoras_cobertura_{args.base}_{stamp}.xlsx")
    result = export_revogadoras_coverage(args.base, output_path, docs_dir_override=args.docs_dir)

    print(f"[OK] base: {args.base}")
    print(f"[OK] revogadoras distintas: {len(result['distinct_rows'])}")
    print(f"[OK] ausentes na pasta fisica: {len(result['missing_physical'])}")
    print(f"[OK] ausentes no chunks: {len(result['missing_chunks'])}")
    print(f"[OK] ausentes no registry: {len(result['missing_registry'])}")
    print(f"[OK] arquivo gerado: {result['output_path']}")


if __name__ == "__main__":
    main()
