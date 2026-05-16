import argparse
import csv
import json
import os
import re
import shutil
import tempfile
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl


def _strip_accents(text: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(text or ""))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_doc_name(name: Any) -> str:
    value = str(name or "").strip().replace("\\", "/").split("/")[-1].lower()
    if value.endswith(".pdf"):
        value = value[:-4]
    value = _strip_accents(value)
    return re.sub(r"[^a-z0-9]+", "", value)


def _coerce_iso_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    # Excel strings commonly arrive already as ISO dates.
    try:
        return datetime.fromisoformat(text[:10]).date().isoformat()
    except ValueError:
        pass

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    raise ValueError(f"Data invalida na planilha: {value!r}")


def _header_key(value: Any) -> str:
    value = _strip_accents(str(value or "").strip().lower())
    return re.sub(r"[^a-z0-9]+", "_", value).strip("_")


def _pick(row: Dict[str, Any], candidates: Iterable[str]) -> Any:
    for name in candidates:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return None


def load_sheet_updates(xlsx_path: Path, sheet_name: Optional[str]) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise RuntimeError("Planilha vazia.") from exc

    headers = [_header_key(h) for h in raw_headers]
    updates: Dict[str, Dict[str, Any]] = {}
    duplicate_docs: List[str] = []

    for row_number, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        raw_doc = _pick(row, ("doc_resolvido", "arquivo", "doc", "doc_name", "source_file_name"))
        if not raw_doc:
            continue

        data_publicacao = _coerce_iso_date(
            _pick(row, ("data_publicacao", "publicacao", "dt_publicacao"))
        )
        data_inicio = _coerce_iso_date(
            _pick(row, ("data_inicio", "data_inicio_vigencia", "inicio_vigencia", "dt_inicio"))
        )

        if not data_publicacao and not data_inicio:
            continue

        doc_name = str(raw_doc).strip()
        key = _normalize_doc_name(doc_name)
        if not key:
            continue
        if key in updates:
            duplicate_docs.append(doc_name)

        updates[key] = {
            "row_number": row_number,
            "doc_name": doc_name,
            "data_publicacao": data_publicacao,
            "data_inicio_vigencia": data_inicio,
        }

    return updates, duplicate_docs


def _iter_chunks(chunks_path: Path):
    with chunks_path.open("r", encoding="utf-8") as handle:
        for line_no, raw in enumerate(handle, start=1):
            raw = raw.strip()
            if not raw:
                continue
            try:
                yield line_no, json.loads(raw)
            except json.JSONDecodeError as exc:
                raise ValueError(f"JSON invalido em {chunks_path} linha {line_no}: {exc}") from exc


def _backup_file(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.with_suffix(path.suffix + f".bak_{stamp}_vigencia_sheet")
    shutil.copy2(path, backup)
    return backup


def _atomic_write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w",
        delete=False,
        dir=path.parent,
        encoding="utf-8",
        newline="\n",
    ) as tmp:
        for row in rows:
            tmp.write(json.dumps(row, ensure_ascii=False) + "\n")
        tmp_path = Path(tmp.name)

    try:
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def apply_updates(
    chunks_path: Path,
    updates: Dict[str, Dict[str, Any]],
    dry_run: bool,
    create_backup: bool,
    changed_output_path: Optional[Path],
) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]], Optional[Path]]:
    enriched_rows: List[Dict[str, Any]] = []
    changed_rows: List[Dict[str, Any]] = []
    report_rows: List[Dict[str, Any]] = []
    matched_update_keys = set()
    changed_docs = set()
    matched_docs = set()
    chunks_changed = 0
    pub_changes = 0
    start_changes = 0

    for line_no, chunk in _iter_chunks(chunks_path):
        doc_name = chunk.get("doc_name") or chunk.get("doc")
        update = updates.get(_normalize_doc_name(doc_name))
        before_pub = chunk.get("data_publicacao")
        before_start = chunk.get("data_inicio_vigencia")

        if update:
            matched_update_keys.add(_normalize_doc_name(update["doc_name"]))
            matched_docs.add(str(doc_name))
            changed = False

            new_pub = update.get("data_publicacao")
            if new_pub and chunk.get("data_publicacao") != new_pub:
                chunk["data_publicacao"] = new_pub
                pub_changes += 1
                changed = True

            new_start = update.get("data_inicio_vigencia")
            if new_start and chunk.get("data_inicio_vigencia") != new_start:
                chunk["data_inicio_vigencia"] = new_start
                start_changes += 1
                changed = True

            if changed:
                chunks_changed += 1
                changed_docs.add(str(doc_name))
                changed_rows.append(chunk)
                report_rows.append(
                    {
                        "line_no": line_no,
                        "doc_name": doc_name,
                        "chunk_id": chunk.get("chunk_id"),
                        "data_publicacao_antes": before_pub,
                        "data_publicacao_depois": chunk.get("data_publicacao"),
                        "data_inicio_vigencia_antes": before_start,
                        "data_inicio_vigencia_depois": chunk.get("data_inicio_vigencia"),
                    }
                )

        enriched_rows.append(chunk)

    backup_path = None
    if not dry_run:
        if create_backup:
            backup_path = _backup_file(chunks_path)
        _atomic_write_jsonl(chunks_path, enriched_rows)

    if changed_output_path:
        changed_output_path.parent.mkdir(parents=True, exist_ok=True)
        _atomic_write_jsonl(changed_output_path, changed_rows)

    missing_updates = [
        update["doc_name"]
        for key, update in sorted(updates.items(), key=lambda item: item[1]["doc_name"].lower())
        if key not in matched_update_keys
    ]

    summary = {
        "dry_run": dry_run,
        "rows_planilha_validas": len(updates),
        "docs_planilha_encontrados": len(matched_update_keys),
        "docs_planilha_nao_encontrados": len(missing_updates),
        "docs_com_chunks_alterados": len(changed_docs),
        "chunks_alterados": chunks_changed,
        "alteracoes_data_publicacao_em_chunks": pub_changes,
        "alteracoes_data_inicio_vigencia_em_chunks": start_changes,
        "backup": str(backup_path) if backup_path else "",
    }
    return summary, report_rows, missing_updates, backup_path


def write_report(report_path: Path, report_rows: List[Dict[str, Any]]) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "line_no",
        "doc_name",
        "chunk_id",
        "data_publicacao_antes",
        "data_publicacao_depois",
        "data_inicio_vigencia_antes",
        "data_inicio_vigencia_depois",
    ]
    with report_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report_rows)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Aplica datas revisadas de uma planilha aos metadados do chunks.jsonl."
    )
    parser.add_argument("--xlsx", required=True, type=Path, help="Planilha XLSX revisada.")
    parser.add_argument("--chunks", required=True, type=Path, help="Caminho do chunks.jsonl.")
    parser.add_argument("--sheet", help="Nome da aba. Padrao: primeira aba.")
    parser.add_argument("--report", type=Path, help="CSV de conferencia das alteracoes.")
    parser.add_argument("--changed-output", type=Path, help="JSONL apenas com chunks alterados.")
    parser.add_argument("--dry-run", action="store_true", help="Simula sem sobrescrever chunks.jsonl.")
    parser.add_argument("--no-backup", action="store_true", help="Nao cria backup do chunks.jsonl.")
    return parser.parse_args()


def main():
    args = parse_args()
    updates, duplicate_docs = load_sheet_updates(args.xlsx, args.sheet)
    if not updates:
        raise RuntimeError("Nenhuma linha valida encontrada na planilha.")

    summary, report_rows, missing_updates, _backup_path = apply_updates(
        chunks_path=args.chunks,
        updates=updates,
        dry_run=bool(args.dry_run),
        create_backup=not args.no_backup,
        changed_output_path=args.changed_output,
    )

    if args.report:
        write_report(args.report, report_rows)

    print(json.dumps(summary, indent=2, ensure_ascii=False))
    if duplicate_docs:
        print(f"[WARN] documentos duplicados na planilha: {len(duplicate_docs)}")
    if missing_updates:
        print(f"[WARN] documentos da planilha nao encontrados: {len(missing_updates)}")
        print(json.dumps(missing_updates[:20], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
