import argparse
import re
import sys
from pathlib import Path
from typing import List, Optional, Tuple

import fitz  # pymupdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.processor import extract_text


def _normalize_spaces(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _first_page_text(pdf_path: Path) -> str:
    with fitz.open(pdf_path) as doc:
        if len(doc) == 0:
            return ""
        return extract_text(doc[0]) or ""


def _extract_header_revocation_parser(text: str) -> Optional[str]:
    if not text:
        return None

    lines = [_normalize_spaces(line) for line in str(text).splitlines() if str(line).strip()]
    if not lines:
        return None

    # A revogacao costuma aparecer logo depois do cabecalho, nas primeiras
    # linhas da primeira pagina.
    header_window = lines[:20]

    patterns = [
        re.compile(r"^\s*(?:parcialmente\s+)?revogad[ao]s?\s+pela\s+.+$", re.IGNORECASE),
    ]

    for line in header_window:
        normalized = _normalize_spaces(line)
        for pattern in patterns:
            if pattern.search(normalized):
                return normalized

    # Fallback: alguns PDFs embutem o cabecalho em uma linha corrida.
    prefix = _normalize_spaces(" ".join(lines[:10]))
    m = re.search(
        r"^\s*(?:parcialmente\s+)?revogad[ao]s?\s+pela\s+.{6,220}$",
        prefix,
        flags=re.IGNORECASE,
    )
    if m:
        return _normalize_spaces(m.group(0))

    return None


def _build_rows(pdf_dir: Path) -> List[Tuple[str, str]]:
    rows: List[Tuple[str, str]] = []
    pdfs = sorted(pdf_dir.glob("*.pdf"))
    total = len(pdfs)

    for idx, pdf_path in enumerate(pdfs, start=1):
        print(f"[{idx}/{total}] {pdf_path.name}")
        try:
            text = _first_page_text(pdf_path)
        except Exception as exc:
            print(f"[WARN] Falha ao ler {pdf_path.name}: {exc}")
            continue

        parser = _extract_header_revocation_parser(text)
        if parser:
            rows.append((pdf_path.name, parser))

    return rows


def _write_excel(rows: List[Tuple[str, str]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revogacao Cabecalho"

    ws.append(["documento", "parser"])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for doc_name, parser in rows:
        ws.append([doc_name, parser])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {1: 72, 2: 88}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extrai a linha de revogacao logo apos o cabecalho da primeira pagina "
            "e exporta os resultados para Excel."
        )
    )
    parser.add_argument(
        "--pdf-dir",
        required=True,
        help="Pasta com os PDFs da base.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Caminho do arquivo .xlsx de saida.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    pdf_dir = Path(args.pdf_dir)
    output_path = Path(args.output)

    if not pdf_dir.exists():
        raise FileNotFoundError(f"Pasta nao encontrada: {pdf_dir}")

    rows = _build_rows(pdf_dir)
    _write_excel(rows, output_path)
    print(f"[OK] linhas encontradas: {len(rows)}")
    print(f"[OK] excel gerado: {output_path}")


if __name__ == "__main__":
    main()
