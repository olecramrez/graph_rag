import streamlit as st
from pathlib import Path
import sys
import subprocess
import os
import json
import re
import unicodedata
from collections import OrderedDict
from datetime import datetime
from xml.sax.saxutils import escape
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    HRFlowable,
    KeepTogether,
    Preformatted,
)

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image, ImageChops
    PIL_AVAILABLE = True
except ImportError:
    Image = None
    ImageChops = None
    PIL_AVAILABLE = False



# =====================================================
# GARANTIR PRIORIDADE DO PROJETO NO SYS.PATH
# =====================================================

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from src.config import (
    DEFAULT_SHARED_BASE_ROOT,
    get_available_bases,
    get_chunks_path,
    get_sqlite_files,
    get_raw_dir,
    USER_ENV,
    get_user_exports_dir,
    get_user_root,
    get_shared_base_root,
    get_documentos_root,
    set_shared_base_root,
    reset_shared_base_root,
)

from src.rag_pipeline import ask, ask_multi_base, DEFAULT_PROMPT_TEMPLATE
from src.cnpj_query import answer_cnpj_query, is_cnpj_query
from src.anm_query import answer_anm_query
from src.universal_sqlite_query import answer_universal_sqlite_query
from src.sqlite_schema_library import load_or_build_schema_profile
from src.sql_agent import compact_schema_for_prompt
from src.command_router import parse_route_command
from src.lia_client import LIAClientError, chat_completion
from src.safe_jsonl import load_valid_jsonl
from src.conversation_memory import (
    append_conversation_turn,
    append_memory_summary,
    answer_conversation_meta_query,
    build_memory_context,
    clear_memory,
    memory_stats,
)
from src.positional_index import load_positional_index, search
from src.lexical_filter import (
    parse_boolean_query,
    docs_from_lexical_query_parts,
    doc_name_matches_lexical_query,
    strip_lexical_syntax,
)

FINAL_LLM_MODELS = ["gpt-5.4", "gpt-5.3-chat", "gpt-4.1"]
SAVED_PROMPTS_FILENAME = "prompts_salvos.json"
SAVED_BASE_ROOTS_FILENAME = "base_roots_salvas.json"

BUILTIN_PROMPT_TEMPLATE_OPTIONS = OrderedDict(
    [
        ("Juridico", DEFAULT_PROMPT_TEMPLATE),
        (
            "Auditoria",
            """
Voce e um assistente de auditoria governamental especializado em analise tecnica de documentos normativos e evidencias.

Data atual para referencia temporal: {data_referencia_iso} ({data_referencia_br})

Responda com base exclusivamente nas evidencias fornecidas no contexto, sem extrapolacoes.

REGRAS OBRIGATORIAS:
1. Utilize apenas informacoes presentes no contexto.
2. Cite documento e pagina para cada achado, criterio, evidencia ou conclusao material.
3. Separe claramente fatos comprovados, lacunas de evidencia, riscos e pontos que exigem validacao manual.
4. Quando houver metadados normativos, considere vigencia, revogacao e aplicabilidade temporal.
5. Nao atribua irregularidade, dolo, culpa ou responsabilidade sem suporte textual suficiente no contexto.
6. Caso a informacao nao esteja no contexto, diga explicitamente que nao foi encontrada.
7. Ao citar artigo/inciso/paragrafo/alinea, use a numeracao literal do contexto.

Estruture a resposta em:
1. Introducao
2. Escopo e criterio de auditoria
3. Analise das evidencias
4. Achados, riscos e lacunas
5. Conclusao

Mapa de referencias normativas extraidas do contexto:
{referencias_normativas_contexto}

Pergunta:
{query}

Contexto:
{context}
""",
        ),
        (
            "Resumo",
            """
Voce e um assistente tecnico especializado em produzir resumos claros, fieis e estruturados de documentos normativos e administrativos.

Data atual para referencia temporal: {data_referencia_iso} ({data_referencia_br})

Responda com base exclusivamente nas evidencias fornecidas no contexto, sem extrapolacoes.

REGRAS OBRIGATORIAS:
1. Preserve o sentido tecnico e juridico do contexto.
2. Cite documento e pagina ao mencionar informacoes materiais.
3. Destaque datas, prazos, competencias, obrigacoes, proibicoes e excecoes relevantes.
4. Quando houver metadados normativos, indique vigencia, revogacao ou incerteza aplicavel.
5. Caso a informacao nao esteja no contexto, diga explicitamente que nao foi encontrada.
6. Nao invente detalhes, exemplos ou conclusoes ausentes.

Estruture a resposta em:
1. Introducao
2. Analise
3. Pontos principais
4. Conclusao

Mapa de referencias normativas extraidas do contexto:
{referencias_normativas_contexto}

Pergunta:
{query}

Contexto:
{context}
""",
        ),
    ]
)


def get_saved_prompts_path():
    return get_user_root() / SAVED_PROMPTS_FILENAME


def load_saved_prompt_templates():
    path = get_saved_prompts_path()
    if not path.exists():
        return OrderedDict()

    try:
        with path.open("r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return OrderedDict()

    if not isinstance(payload, dict):
        return OrderedDict()

    prompts = payload.get("prompts", payload)
    if not isinstance(prompts, dict):
        return OrderedDict()

    loaded = OrderedDict()
    for name, prompt in prompts.items():
        clean_name = str(name or "").strip()
        if clean_name and isinstance(prompt, str):
            loaded[clean_name] = prompt
    return loaded


def save_prompt_template(name, prompt):
    clean_name = str(name or "").strip()
    if not clean_name:
        raise ValueError("Informe um nome para salvar o prompt.")

    prompts = load_saved_prompt_templates()
    prompts[clean_name] = str(prompt or "")

    path = get_saved_prompts_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump({"prompts": prompts}, f, ensure_ascii=False, indent=2)

    return clean_name, path


def get_saved_base_roots_path():
    user_name = os.getenv("USERNAME") or os.getenv("USER") or Path.home().name
    user_dir = DEFAULT_SHARED_BASE_ROOT / "users" / user_name
    user_dir.mkdir(parents=True, exist_ok=True)
    return user_dir / SAVED_BASE_ROOTS_FILENAME


def _read_saved_base_roots_file(path):
    try:
        with path.open("r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception:
        return {"roots": OrderedDict(), "default": ""}

    if not isinstance(payload, dict):
        return {"roots": OrderedDict(), "default": ""}

    raw_roots = payload.get("roots", {})
    roots = OrderedDict()

    if isinstance(raw_roots, dict):
        for name, root in raw_roots.items():
            clean_name = str(name or "").strip()
            clean_root = str(root or "").strip()
            if clean_name and clean_root:
                roots[clean_name] = clean_root

    default_name = str(payload.get("default", "") or "").strip()
    if default_name not in roots:
        default_name = ""

    return {"roots": roots, "default": default_name}


def load_saved_base_roots():
    path = get_saved_base_roots_path()
    if not path.exists():
        legacy_path = get_user_root() / SAVED_BASE_ROOTS_FILENAME
        if legacy_path.exists() and legacy_path.resolve() != path.resolve():
            payload = _read_saved_base_roots_file(legacy_path)
            if payload["roots"]:
                path.parent.mkdir(parents=True, exist_ok=True)
                with path.open("w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
                return payload
        return {"roots": OrderedDict(), "default": ""}

    return _read_saved_base_roots_file(path)


def save_base_root_entry(name, root, make_default=False):
    clean_name = str(name or "").strip()
    clean_root = str(root or "").strip()

    if not clean_name:
        raise ValueError("Informe um nome para salvar o caminho.")

    if not clean_root:
        raise ValueError("Informe um caminho valido.")

    payload = load_saved_base_roots()
    roots = payload["roots"]
    roots[clean_name] = clean_root

    default_name = payload.get("default", "")
    if make_default or not default_name:
        default_name = clean_name

    path = get_saved_base_roots_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(
            {"default": default_name, "roots": roots},
            f,
            ensure_ascii=False,
            indent=2,
        )

    return clean_name, path


def get_default_saved_base_root():
    payload = load_saved_base_roots()
    default_name = payload.get("default", "")
    if not default_name:
        return ""
    return payload["roots"].get(default_name, "")


def validate_base_root_path(root_value):
    root_path = Path(str(root_value or "").strip())
    if not str(root_path):
        return None, "Informe um caminho valido."

    documentos_path = root_path / "documentos"
    data_path = root_path / "data"

    if not root_path.exists():
        return None, "A pasta informada nao existe."
    if not documentos_path.exists():
        return None, "A pasta informada nao contem a subpasta `documentos`."
    if not data_path.exists():
        return None, "A pasta informada nao contem a subpasta `data`."

    return root_path, ""


def apply_base_root_choice():
    choice = st.session_state.get("base_root_choice") or "Padrão"

    if choice == "Padrão":
        st.session_state.custom_base_root = ""
        reset_shared_base_root()
        clear_runtime_caches()
        st.session_state.base_root_feedback = "Caminho 'Padrão' aplicado."
        return

    saved_roots = load_saved_base_roots()["roots"]
    root_value = saved_roots.get(choice, "")
    root_path, error = validate_base_root_path(root_value)

    if error:
        st.session_state.base_root_error = error
        return

    st.session_state.custom_base_root = str(root_path)
    set_shared_base_root(root_path)
    clear_runtime_caches()
    st.session_state.base_root_feedback = f"Caminho '{choice}' aplicado."


def get_prompt_template_options():
    options = OrderedDict(BUILTIN_PROMPT_TEMPLATE_OPTIONS)
    options.update(load_saved_prompt_templates())
    return options


def apply_prompt_template_selection():
    selected = st.session_state.get("prompt_template_choice", "Juridico")
    st.session_state.prompt_custom = get_prompt_template_options().get(
        selected,
        DEFAULT_PROMPT_TEMPLATE,
    )
def recompute_content_refinement():

    lexical_base_docs = st.session_state.get("lexical_base_docs")
    lexical_base_docs_by_base = st.session_state.get("lexical_base_docs_by_base")

    if lexical_base_docs_by_base is not None:
        filtered_by_base = {
            base_name: list(docs or [])
            for base_name, docs in lexical_base_docs_by_base.items()
        }

        for filtro in st.session_state.get("content_filters", []):

            if not filtro.get("active", True):
                continue

            docs_by_base = filtro.get("docs_by_base")

            if docs_by_base is not None:
                for base_name in list(filtered_by_base.keys()):
                    docs_match = set(docs_by_base.get(base_name, []))
                    filtered_by_base[base_name] = [
                        doc for doc in filtered_by_base[base_name]
                        if doc in docs_match
                    ]
            else:
                docs_match = set(filtro.get("docs", []))
                for base_name in list(filtered_by_base.keys()):
                    filtered_by_base[base_name] = [
                        doc for doc in filtered_by_base[base_name]
                        if doc in docs_match
                    ]

        st.session_state.filtered_docs_by_base = filtered_by_base
        current_base = st.session_state.get("last_base")
        if current_base in filtered_by_base:
            st.session_state.filtered_docs = filtered_by_base[current_base]
        elif filtered_by_base:
            st.session_state.filtered_docs = next(iter(filtered_by_base.values()))
        else:
            st.session_state.filtered_docs = []
        return

    if lexical_base_docs is None:
        st.session_state.filtered_docs = None
        st.session_state.filtered_docs_by_base = None
        return

    docs_filtrados = list(lexical_base_docs)

    for filtro in st.session_state.get("content_filters", []):

        if not filtro.get("active", True):
            continue

        docs_match = set(filtro.get("docs", []))

        docs_filtrados = [
            doc for doc in docs_filtrados
            if doc in docs_match
        ]

    st.session_state.filtered_docs = docs_filtrados
    st.session_state.filtered_docs_by_base = None


def get_filter_base_names():

    if st.session_state.get("multi_base_enabled"):
        return [
            str(base_name).strip()
            for base_name in st.session_state.get("multi_selected_bases", [])
            if str(base_name).strip()
        ]

    base_name = st.session_state.get("last_base")
    return [base_name] if base_name else []


def get_allowed_docs_for_query():

    if st.session_state.get("multi_base_enabled"):
        filtered_by_base = st.session_state.get("filtered_docs_by_base")
        if filtered_by_base is None:
            return None

        return {
            base_name: filtered_by_base.get(base_name, [])
            for base_name in get_filter_base_names()
        }

    return st.session_state.get("filtered_docs")


def parse_docs_directive(query):
    text = str(query or "")
    match = re.search(r"@docs\b", text, flags=re.IGNORECASE)
    if not match:
        return text, []

    docs = []
    pos = match.end()
    length = len(text)

    while pos < length:
        while pos < length and (text[pos].isspace() or text[pos] in ",;"):
            pos += 1

        if pos >= length:
            break

        char = text[pos]
        if char in {'"', "'", "`"}:
            end = text.find(char, pos + 1)
            if end == -1:
                break
            doc_name = text[pos + 1:end].strip()
            if doc_name:
                docs.append(doc_name)
            pos = end + 1
            continue

        token_match = re.match(r"[^\s,;]+", text[pos:])
        if not token_match:
            break

        token = token_match.group(0).strip()
        if ".pdf" not in token.lower():
            break

        docs.append(token)
        pos += len(token)

    if not docs:
        return text, []

    while pos < length and text[pos].isspace():
        pos += 1
    if pos < length and text[pos] in ".:;-":
        pos += 1

    cleaned = (text[:match.start()] + " " + text[pos:]).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned, docs


def _normalize_doc_lookup_key(value):
    normalized = unicodedata.normalize("NFKD", str(value or "").casefold())
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", normalized).strip()


def _available_pdf_names(base_name):
    base_dir = get_documentos_root() / base_name
    return sorted(pdf.name for pdf in base_dir.glob("*.pdf") if pdf.is_file())


def _resolve_requested_docs_for_base(requested_docs, base_name, current_allowed=None):
    available_docs = _available_pdf_names(base_name)
    if current_allowed is not None:
        allowed_set = set(current_allowed or [])
        available_docs = [doc for doc in available_docs if doc in allowed_set]

    exact_lookup = {
        _normalize_doc_lookup_key(doc): doc
        for doc in available_docs
    }

    resolved = []
    missing = []

    for requested in requested_docs:
        requested_key = _normalize_doc_lookup_key(requested)
        match_doc = exact_lookup.get(requested_key)

        if not match_doc:
            partial_matches = [
                doc for doc in available_docs
                if requested_key and requested_key in _normalize_doc_lookup_key(doc)
            ]
            if len(partial_matches) == 1:
                match_doc = partial_matches[0]

        if match_doc and match_doc not in resolved:
            resolved.append(match_doc)
        elif not match_doc:
            missing.append(requested)

    return resolved, missing


def build_docs_directive_scope(requested_docs):
    current_allowed = get_allowed_docs_for_query()

    if not requested_docs:
        return current_allowed, {}, []

    if st.session_state.get("multi_base_enabled"):
        resolved_by_base = {}
        missing = set(requested_docs)

        for base_name in get_filter_base_names():
            base_allowed = (
                current_allowed.get(base_name)
                if isinstance(current_allowed, dict)
                else current_allowed
            )
            resolved, _ = _resolve_requested_docs_for_base(
                requested_docs,
                base_name,
                current_allowed=base_allowed,
            )
            resolved_by_base[base_name] = resolved
            for doc in resolved:
                requested_key = _normalize_doc_lookup_key(doc)
                missing = {
                    item for item in missing
                    if _normalize_doc_lookup_key(item) != requested_key
                }

        return resolved_by_base, resolved_by_base, sorted(missing)

    base_name = st.session_state.get("last_base")
    if not base_name:
        return [], {}, list(requested_docs)

    resolved, missing = _resolve_requested_docs_for_base(
        requested_docs,
        base_name,
        current_allowed=current_allowed,
    )
    return resolved, {base_name: resolved}, missing


def format_docs_directive_notice(resolved_by_base, missing_docs=None):
    lines = ["Consulta limitada por @docs aos documentos encontrados:"]
    shown_any = False

    for base_name, docs in (resolved_by_base or {}).items():
        for doc in docs or []:
            shown_any = True
            if st.session_state.get("multi_base_enabled"):
                lines.append(f"- {base_name}: {doc}")
            else:
                lines.append(f"- {doc}")

    if not shown_any:
        lines.append("- Nenhum documento solicitado foi encontrado no escopo atual.")

    for doc in missing_docs or []:
        lines.append(f"- Nao encontrado: {doc}")

    return "\n".join(lines)


def format_docs_directive_timing_notice(resolved_by_base, missing_docs=None):
    found = []
    for base_name, docs in (resolved_by_base or {}).items():
        for doc in docs or []:
            if st.session_state.get("multi_base_enabled"):
                found.append(f"{base_name}: {doc}")
            else:
                found.append(doc)

    if not found:
        message = "[INFO] @docs: nenhum documento solicitado foi encontrado no escopo atual."
    else:
        message = "[INFO] @docs limitado a: " + "; ".join(found)

    if missing_docs:
        message += " | nao encontrados: " + "; ".join(str(doc) for doc in missing_docs)

    return message


def get_active_base_label():

    if st.session_state.get("multi_base_enabled"):
        return ", ".join(get_filter_base_names())

    return st.session_state.get("last_base") or ""


def get_total_docs_for_active_scope():

    if st.session_state.get("multi_base_enabled"):
        return sum(
            len(list((get_documentos_root() / base_name).glob("*.pdf")))
            for base_name in get_filter_base_names()
        )

    base_name = st.session_state.get("last_base")
    if not base_name:
        return 0

    return len(list((get_documentos_root() / base_name).glob("*.pdf")))


def get_filtered_docs_total():

    filtered_by_base = st.session_state.get("filtered_docs_by_base")

    if st.session_state.get("multi_base_enabled") and filtered_by_base is not None:
        return sum(
            len(filtered_by_base.get(base_name, []))
            for base_name in get_filter_base_names()
        )

    filtered_docs = st.session_state.get("filtered_docs")
    return len(filtered_docs or [])


def get_sidebar_doc_entries():

    if st.session_state.get("multi_base_enabled"):
        base_names = get_filter_base_names()
        filtered_by_base = st.session_state.get("filtered_docs_by_base")
        entries = []

        for base_name in base_names:
            base_dir = get_documentos_root() / base_name

            if filtered_by_base is not None:
                docs = filtered_by_base.get(base_name, [])
                entries.extend((base_name, base_dir / doc) for doc in docs)
            else:
                entries.extend((base_name, pdf) for pdf in base_dir.glob("*.pdf"))

        return entries

    base_name = st.session_state.get("last_base") or ""
    base_dir = get_documentos_root() / base_name

    if st.session_state.get("filtered_docs") is not None:
        return [
            (base_name, base_dir / doc)
            for doc in st.session_state.filtered_docs
        ]

    return [(base_name, pdf) for pdf in base_dir.glob("*.pdf")]


def extract_terms_from_query(query):

    if not query:
        return []

    terms = []
    seen = set()

    for phrase in re.findall(r'"([^"]+)"', query):

        normalized = phrase.strip().lower()

        if normalized and normalized not in seen:
            seen.add(normalized)
            terms.append(normalized)

    query_clean = re.sub(r'"[^"]+"', " ", query)

    for token in re.findall(r"\w+", query_clean.lower()):

        if token in {"and", "or", "not", "near"}:
            continue

        if token.isdigit():
            continue

        if token not in seen:
            seen.add(token)
            terms.append(token)

    return terms


def count_term_occurrences(text, term):

    if not text or not term:
        return 0

    term_norm = term.strip().lower()

    if not term_norm:
        return 0

    text_norm = text.lower()

    if " " in term_norm:
        return len(re.findall(re.escape(term_norm), text_norm))

    pattern = rf"\b{re.escape(term_norm)}\b"
    return len(re.findall(pattern, text_norm))


def collect_active_terms():

    terms = []
    seen = set()

    for term in extract_terms_from_query(st.session_state.get("termos_busca", "")):

        if term not in seen:
            seen.add(term)
            terms.append(term)

    for filtro in st.session_state.get("content_filters", []):

        if not filtro.get("active", True):
            continue

        for term in extract_terms_from_query(filtro.get("term", "")):

            if term not in seen:
                seen.add(term)
                terms.append(term)

    return terms


def get_query_total_occurrences(query_text, docs, doc_text_lookup):

    query_terms = extract_terms_from_query(query_text)

    if not query_terms and query_text.strip():
        query_terms = [query_text.strip().lower()]

    total = 0

    for doc in docs:

        doc_text = doc_text_lookup.get(doc, "")

        for term in query_terms:
            total += count_term_occurrences(doc_text, term)

    return total


def describe_lexical_filter(query_text):
    terms, operator, exclude = parse_boolean_query(query_text)
    include_terms = [
        term for term in (strip_lexical_syntax(t) for t in terms)
        if term
    ]
    exclude_terms = [
        term for term in (strip_lexical_syntax(t) for t in exclude)
        if term
    ]

    parts = []
    if include_terms:
        parts.append(f"Incluir ({operator}): {', '.join(include_terms)}")
    if exclude_terms:
        parts.append(f"Excluir: {', '.join(exclude_terms)}")

    return " | ".join(parts) or str(query_text or "").strip()


def is_negative_only_filter(query_text):
    terms, _, exclude = parse_boolean_query(query_text)
    return bool(exclude and not terms)


def fix_mojibake_text(text):

    if not text:
        return text

    if "" not in text and "" not in text:
        return text

    try:
        return text.encode("latin1").decode("utf-8")
    except UnicodeError:
        return text


def normalize_doc_key(name):

    if not name:
        return ""

    name = fix_mojibake_text(name)
    normalized = unicodedata.normalize("NFKD", name)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", ascii_only.lower())


def resolve_doc_name(doc_name, available_docs_set, available_docs_by_key):

    if doc_name in available_docs_set:
        return doc_name

    doc_fixed = fix_mojibake_text(doc_name)

    if doc_fixed in available_docs_set:
        return doc_fixed

    key_original = normalize_doc_key(doc_name)
    key_fixed = normalize_doc_key(doc_fixed)

    if key_original in available_docs_by_key:
        return available_docs_by_key[key_original]

    if key_fixed in available_docs_by_key:
        return available_docs_by_key[key_fixed]

    return doc_name


def build_doc_name_maps(pdfs):

    available_doc_names = [pdf.name for pdf in pdfs]
    available_docs_set = set(available_doc_names)
    available_docs_by_key = {}

    for name in available_doc_names:
        key = normalize_doc_key(name)
        if key and key not in available_docs_by_key:
            available_docs_by_key[key] = name

    return available_docs_set, available_docs_by_key


def build_cited_docs_from_results(results, max_items=12):

    cited = OrderedDict()

    for score, chunk in (results or [])[:max_items]:

        doc = chunk.get("doc") or chunk.get("doc_name")

        if not doc:
            continue

        if doc not in cited:
            cited[doc] = {
                "documento": doc,
                "paginas": [],
                "score_max": score,
                "tipo_norma": chunk.get("tipo_norma") or "",
                "numero_norma": chunk.get("numero_norma") or "",
                "ano_norma": chunk.get("ano_norma") or "",
                "data_publicacao": chunk.get("data_publicacao") or "",
                "data_inicio_vigencia": chunk.get("data_inicio_vigencia") or "",
                "data_fim_vigencia": chunk.get("data_fim_vigencia") or "",
                "status_normativo": chunk.get("status_normativo") or "",
                "tipo_revogacao": chunk.get("tipo_revogacao") or "",
                "revogado_por": chunk.get("revogado_por") or "",
            }

        page = chunk.get("page")
        if page not in (None, "") and page not in cited[doc]["paginas"]:
            cited[doc]["paginas"].append(page)

        if score > cited[doc]["score_max"]:
            cited[doc]["score_max"] = score

    return list(cited.values())


def build_cited_docs_from_evidence(evidence_text):

    cited = OrderedDict()
    pattern = re.compile(
        r"^###\s+(?P<doc>.+?)\s+\|\s+Pagina\s+(?P<page>.*?)\s+\|\s+Score\s+(?P<score>[0-9.,-]+)",
        re.IGNORECASE | re.MULTILINE,
    )

    for match in pattern.finditer(evidence_text or ""):
        doc = match.group("doc").strip()
        if not doc:
            continue
        if doc not in cited:
            cited[doc] = {
                "documento": doc,
                "paginas": [],
                "score_max": "",
                "tipo_norma": "",
                "numero_norma": "",
                "ano_norma": "",
                "data_publicacao": "",
                "data_inicio_vigencia": "",
                "data_fim_vigencia": "",
                "status_normativo": "",
                "tipo_revogacao": "",
                "revogado_por": "",
            }
        page = match.group("page").strip()
        if page and page not in cited[doc]["paginas"]:
            cited[doc]["paginas"].append(page)
        score_text = match.group("score").replace(",", ".")
        try:
            score = float(score_text)
        except ValueError:
            score = None
        if score is not None and (
            cited[doc]["score_max"] == "" or score > cited[doc]["score_max"]
        ):
            cited[doc]["score_max"] = score

    return list(cited.values())


def export_cited_docs_excel(
    item,
    response_index,
    base,
    documentos_dir,
    available_docs_set,
    available_docs_by_key,
):

    cited_docs = item.get("cited_docs") or build_cited_docs_from_evidence(
        item.get("evidence", "")
    )

    if not cited_docs:
        raise ValueError("Nenhum documento citado foi encontrado nessa resposta.")

    export_dir = get_user_exports_dir()
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_path = (
        export_dir
        / f"{base}_documentos_citados_resposta_{response_index + 1}_{timestamp}.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Documentos citados"

    headers = [
        "documento",
        "paginas_citadas",
        "score_max",
        "tipo_norma",
        "numero_norma",
        "ano_norma",
        "data_publicacao",
        "data_inicio_vigencia",
        "data_fim_vigencia",
        "status_normativo",
        "tipo_revogacao",
        "revogado_por",
    ]
    sheet.append(headers)

    for cited in cited_docs:
        resolved_doc = resolve_doc_name(
            cited.get("documento", ""),
            available_docs_set,
            available_docs_by_key,
        )
        paginas = cited.get("paginas") or []
        sheet.append([
            resolved_doc,
            ", ".join(str(page) for page in paginas),
            cited.get("score_max", ""),
            cited.get("tipo_norma", ""),
            cited.get("numero_norma", ""),
            cited.get("ano_norma", ""),
            cited.get("data_publicacao", ""),
            cited.get("data_inicio_vigencia", ""),
            cited.get("data_fim_vigencia", ""),
            cited.get("status_normativo", ""),
            cited.get("tipo_revogacao", ""),
            cited.get("revogado_por", ""),
        ])

        doc_cell = sheet.cell(row=sheet.max_row, column=1)
        doc_cell.hyperlink = str(documentos_dir / resolved_doc)
        doc_cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 80,
        "B": 18,
        "C": 12,
        "D": 16,
        "E": 14,
        "F": 12,
        "G": 18,
        "H": 22,
        "I": 18,
        "J": 20,
        "K": 18,
        "L": 80,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    info = workbook.create_sheet("consulta")
    info.append(["campo", "valor"])
    info.append(["base", base])
    info.append(["pergunta", item.get("question", "")])
    info.append(["data_hora_resposta", item.get("timestamp", "")])
    info.append(["modo", item.get("modo", "")])
    info.append(["total_documentos_citados", len(cited_docs)])
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 120

    workbook.save(export_path)
    return export_path


def read_env_var(path, key):

    if not path.exists():
        return ""

    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()

        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue

        current_key, value = line.split("=", 1)

        if current_key.strip() == key:
            return value.strip().strip('"').strip("'")

    return ""


def upsert_env_var(path, key, value):

    path.parent.mkdir(parents=True, exist_ok=True)
    lines = []

    if path.exists():
        lines = path.read_text(encoding="utf-8").splitlines()

    updated = False
    out_lines = []

    for line in lines:
        stripped = line.strip()

        if (
            not updated
            and stripped
            and not stripped.startswith("#")
            and "=" in line
            and line.split("=", 1)[0].strip() == key
        ):
            out_lines.append(f"{key}={value}")
            updated = True
            continue

        out_lines.append(line)

    if not updated:
        if out_lines and out_lines[-1].strip():
            out_lines.append("")
        out_lines.append(f"{key}={value}")

    path.write_text("\n".join(out_lines).rstrip() + "\n", encoding="utf-8")


def get_current_api_key():

    key = (os.getenv("LIA_API_KEY") or "").strip()

    if key:
        return key

    return read_env_var(USER_ENV, "LIA_API_KEY")


def apply_session_base_root():
    root = (st.session_state.get("custom_base_root") or "").strip()
    if root:
        set_shared_base_root(root)


def clear_runtime_caches():
    try:
        st.cache_data.clear()
    except Exception:
        pass
    try:
        st.cache_resource.clear()
    except Exception:
        pass


def split_answer_and_evidence(resposta_completa):
    split_pattern = r"#\s*(?:ðŸ“Ž\s*)?Evid(?:Ãª|e)ncia"
    partes = re.split(split_pattern, resposta_completa, maxsplit=1, flags=re.IGNORECASE)

    if len(partes) == 1:
        legacy_pattern = r"#\s*.*Evid.*ncia"
        partes = re.split(legacy_pattern, resposta_completa, maxsplit=1, flags=re.IGNORECASE)

    resposta_texto = partes[0]
    evidencia_texto = partes[1] if len(partes) > 1 else ""
    return resposta_texto, evidencia_texto


HYBRID_CNPJ_RAG_TERMS = (
    "rag",
    "base",
    "bases",
    "documento",
    "documentos",
    "evidencia",
    "evidencias",
    "norma",
    "normas",
    "normativo",
    "normativos",
    "relacione",
    "relacionar",
    "relacionados",
    "menciona",
    "mencionam",
    "aparece",
    "consta",
    "procure na base",
    "buscar na base",
    "busque na base",
    "verifique na base",
    "cruze",
    "compare",
)


def _plain_ascii(text):
    normalized = unicodedata.normalize("NFKD", str(text or ""))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def is_hybrid_cnpj_rag_query(query):
    normalized = re.sub(r"\s+", " ", _plain_ascii(query).lower()).strip()
    if not normalized:
        return False
    return any(term in normalized for term in HYBRID_CNPJ_RAG_TERMS)


def _clean_cnpj_term(value):
    value = re.sub(r"\*\*|`", "", str(value or "")).strip()
    value = re.sub(r"\s+", " ", value).strip(" :-|")
    if not value:
        return ""
    if value.lower() in {"nao informado", "não informado", "none", "null"}:
        return ""
    return value


def _short_log_text(text, limit=900):
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "..."


def extract_cnpj_terms_for_rag(cnpj_answer_text, max_terms=12):
    terms = []
    seen = set()

    def add_term(value):
        clean = _clean_cnpj_term(value)
        if not clean:
            return
        key = _plain_ascii(clean).lower()
        if key in seen:
            return
        seen.add(key)
        terms.append(clean)

    label_pattern = re.compile(
        r"\*\*(CNPJ|Razao social|Razão social|Nome fantasia|CNAE principal):\*\*\s*([^\n]+)",
        flags=re.IGNORECASE,
    )
    for match in label_pattern.finditer(str(cnpj_answer_text or "")):
        add_term(match.group(2))

    lines = str(cnpj_answer_text or "").splitlines()
    for line in lines:
        if "|" not in line or re.search(r"\|\s*-{3,}", line):
            continue
        cells = [_clean_cnpj_term(cell) for cell in line.strip().strip("|").split("|")]
        for cell in cells:
            if re.fullmatch(r"\d{2}\s*-.*", cell):
                continue
            if cell.lower() in {
                "cnpj",
                "razao social",
                "razão social",
                "nome fantasia",
                "situacao",
                "situação",
                "cnae",
                "uf",
                "municipio",
                "município",
            }:
                continue
            if len(cell) >= 4:
                add_term(cell)

    return terms[:max_terms]


def build_hybrid_rag_query(original_query, cnpj_answer_text):
    terms = extract_cnpj_terms_for_rag(cnpj_answer_text)
    terms_text = "; ".join(terms) if terms else "termos extraidos da consulta CNPJ"
    cnpj_context = re.sub(r"\n{3,}", "\n\n", str(cnpj_answer_text or "")).strip()
    return (
        f"{original_query}\n\n"
        "Use os dados cadastrais CNPJ abaixo como insumo para buscar evidencias nas bases RAG. "
        "Procure mencoes diretas e relacoes com razao social, nome fantasia, CNPJ, CNAE e socios quando existirem. "
        "Nao trate os dados CNPJ como evidencia documental da base RAG; cite documentos/paginas apenas quando vierem do contexto recuperado.\n\n"
        f"Termos prioritarios para busca: {terms_text}\n\n"
        f"Dados CNPJ apurados:\n{cnpj_context[:8000]}"
    )


def build_cnpj_query_from_rag(original_query, rag_answer_text):
    terms = extract_cnpj_terms_for_rag(rag_answer_text, max_terms=16)
    terms_text = "; ".join(terms) if terms else str(original_query or "").strip()
    return (
        f"{original_query}\n\n"
        "Use os achados abaixo vindos da consulta RAG para consultar a base CNPJ/QSA. "
        "Priorize CNPJs, razoes sociais, nomes fantasia e nomes de pessoas/socios identificados.\n\n"
        f"Termos extraidos: {terms_text}\n\n"
        f"Achados RAG:\n{str(rag_answer_text or '').strip()[:8000]}"
    )


def join_hybrid_outputs(cnpj_output, rag_output):
    cnpj_answer, cnpj_evidence = split_answer_and_evidence(cnpj_output)
    rag_answer, rag_evidence = split_answer_and_evidence(rag_output)
    return (
        "# Resposta\n\n"
        "## Consulta CNPJ\n\n"
        f"{cnpj_answer.strip()}\n\n"
        "## Consulta nas bases RAG\n\n"
        f"{rag_answer.strip()}\n\n"
        "---\n\n"
        "# Evidencia\n\n"
        "## Evidencia CNPJ\n\n"
        f"{cnpj_evidence.strip() or 'Sem evidencia CNPJ detalhada.'}\n\n"
        "## Evidencia RAG\n\n"
        f"{rag_evidence.strip() or 'Sem evidencia RAG detalhada.'}"
    )


ALLOWED_SEARCH_ROUTES = {"cnpj_only", "anm_only", "sqlite_only", "rag_only", "cnpj_rag_hybrid", "rag_cnpj_hybrid"}


def fallback_search_route(query):
    qn = _plain_ascii(query).lower()
    if any(term in qn for term in ("depois consulte o cnpj", "depois no cnpj", "apos consultar os documentos", "após consultar os documentos")):
        route = "rag_cnpj_hybrid"
        reason = "A pergunta pede buscar nos documentos primeiro e consultar CNPJ depois."
    elif is_cnpj_query(query):
        if is_hybrid_cnpj_rag_query(query):
            route = "cnpj_rag_hybrid"
            reason = "A pergunta parece citar CNPJ e pedir relacao com documentos/bases."
        else:
            route = "cnpj_only"
            reason = "A pergunta parece ser cadastral de CNPJ."
    else:
        route = "rag_only"
        reason = "A pergunta nao parece exigir consulta CNPJ."

    return {
        "route": route,
        "reason": reason,
        "rag_query": str(query or "").strip(),
        "source": "fallback",
    }


def _extract_json_payload(text):
    raw = str(text or "").strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?", "", raw, flags=re.IGNORECASE).strip()
        raw = re.sub(r"```$", "", raw).strip()

    try:
        return json.loads(raw)
    except Exception:
        start = raw.find("{")
        end = raw.rfind("}")
        if start >= 0 and end > start:
            return json.loads(raw[start:end + 1])
    raise ValueError("Roteador nao retornou JSON valido.")


def sanitize_search_route(payload, query, sqlite_available=False):
    if not isinstance(payload, dict):
        return None

    route = str(payload.get("route") or "").strip()
    if route not in ALLOWED_SEARCH_ROUTES:
        return None
    if route == "sqlite_only" and not sqlite_available:
        route = "rag_only"

    has_cnpj_signal = is_cnpj_query(query)
    if route in {"cnpj_only", "cnpj_rag_hybrid"} and not has_cnpj_signal:
        route = "rag_only"
    if route == "rag_only" and has_cnpj_signal:
        route = fallback_search_route(query)["route"]

    reason = re.sub(r"\s+", " ", str(payload.get("reason") or "")).strip()
    rag_query = re.sub(r"\s+", " ", str(payload.get("rag_query") or "")).strip()

    if not rag_query:
        rag_query = str(query or "").strip()

    return {
        "route": route,
        "reason": reason[:240] or "Rota escolhida pelo roteador LLM.",
        "rag_query": rag_query[:1200],
        "source": "llm",
    }


@st.cache_data(show_spinner=False)
def build_sqlite_router_context(sqlite_db_path, base_name, sqlite_mtime):
    if not sqlite_db_path:
        return ""
    profile, _ = load_or_build_schema_profile(sqlite_db_path, base_name=base_name)
    schema = profile.get("schema") or []
    schema_text = compact_schema_for_prompt(schema, max_tables=12, max_columns=18)
    dictionary_entries = profile.get("dictionary_entries") or 0
    dictionary_note = (
        f"Dicionario incorporado: {dictionary_entries} campo(s)."
        if dictionary_entries
        else "Sem dicionario incorporado."
    )
    return (
        f"SQLite selecionado: {profile.get('db_name')}\n"
        f"{dictionary_note}\n"
        f"Resumo do schema/dicionario:\n{schema_text[:10000]}"
    )


def choose_search_route(
    query,
    llm_model=None,
    progress_callback=None,
    sqlite_db_path=None,
    sqlite_base_name=None,
):
    command_route, command_query = parse_route_command(query)
    if command_route:
        route = {
            "route": f"{command_route}_only",
            "reason": f"Comando explicito @{command_route}.",
            "rag_query": command_query or str(query or "").strip(),
            "source": "command",
        }
        if progress_callback:
            progress_callback(f"[INFO] Roteador por comando: {route['route']} - {route['reason']}")
        return route

    fallback = fallback_search_route(query)
    sqlite_context = ""
    sqlite_available = bool(sqlite_db_path)
    if sqlite_available:
        try:
            sqlite_mtime = Path(sqlite_db_path).stat().st_mtime
            sqlite_context = build_sqlite_router_context(
                str(sqlite_db_path),
                sqlite_base_name or "",
                sqlite_mtime,
            )
        except Exception as exc:
            sqlite_available = False
            if progress_callback:
                progress_callback(f"[WARN] Schema SQLite indisponivel para o router LLM: {exc}")

    prompt = f"""
Escolha a melhor rota de busca para a pergunta do usuario.
Responda somente JSON puro, sem markdown.

Rotas permitidas:
- cnpj_only: usar apenas a base SQLite CNPJ para dados cadastrais, socios, CNAE, situacao, nome fantasia ou razao social.
- sqlite_only: usar apenas o SQLite generico selecionado para a base ativa quando a pergunta puder ser respondida por tabelas/colunas do schema SQLite abaixo, inclusive perguntas sobre dicionario/codigos de campos.
- rag_only: usar apenas as bases RAG/documentos.
- cnpj_rag_hybrid: consultar CNPJ primeiro e depois usar os dados cadastrais como insumo para buscar evidencias nas bases RAG.
- rag_cnpj_hybrid: consultar documentos/RAG primeiro e depois usar empresas, CNPJs ou pessoas encontrados para consultar CNPJ/QSA.

Regras:
- Se a pergunta pede dados cadastrais, QSA, socios, administrador, representante, endereco, situacao cadastral, CNAE, porte, capital social, quadro societario, vinculo/vinculacao societaria ou "todas as informacoes" de uma empresa/pessoa, escolha cnpj_only.
- Se a pergunta pede se uma empresa tem vinculacao/ligacao/conexao com outras empresas por socios, escolha cnpj_only.
- Se a pergunta pede se uma pessoa tem empresa, e socio de empresa ou tem empresa em uma UF, escolha cnpj_only e use QSA/socios, nao busca textual por razao social.
- Se houver SQLite selecionado e a pergunta menciona campos, colunas, codigos, indicadores, tabelas, agentes, anos, totais, rankings, amostras, agregacoes ou termos presentes no schema/dicionario abaixo, escolha sqlite_only.
- Se a pergunta pede "o que significa", "o que e", "traduza", "codigo X" ou "numero X" sobre campo do dicionario, escolha sqlite_only.
- Se a pergunta pede documentos, evidencias, normas, mencoes, relacao com bases ou verificacao no RAG sobre uma empresa/CNPJ, escolha cnpj_rag_hybrid.
- Se a pergunta manda procurar primeiro nos chunks/documentos/bases e depois consultar CNPJ com os achados, escolha rag_cnpj_hybrid.
- Se nao houver tema de CNPJ, escolha rag_only.
- Nao escreva SQL.

Contexto SQLite disponivel:
{sqlite_context or "Nenhum SQLite selecionado para roteamento."}

Formato:
{{
  "route": "cnpj_rag_hybrid",
  "reason": "breve justificativa",
  "rag_query": "consulta sugerida para o RAG, se aplicavel"
}}

Pergunta:
{query}
""".strip()

    try:
        raw = chat_completion(
            [
                {
                    "role": "system",
                    "content": (
                        "Voce e um roteador de consultas. "
                        "Retorne apenas JSON valido com uma rota permitida."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0.0,
            max_retries=2,
            llm_model=llm_model,
        )
        route = sanitize_search_route(
            _extract_json_payload(raw),
            query,
            sqlite_available=sqlite_available,
        )
        if route:
            if progress_callback:
                progress_callback(
                    f"[INFO] Roteador LLM ({llm_model or 'modelo padrao'}): {route['route']} - {route['reason']}"
                )
            return route
        raise ValueError("Rota ausente ou nao permitida.")
    except (LIAClientError, ValueError, json.JSONDecodeError, TypeError) as exc:
        if progress_callback:
            progress_callback(
                f"[WARN] Roteador LLM indisponivel; usando regra local: {exc}"
            )
            progress_callback(
                f"[INFO] Roteador local: {fallback['route']} - {fallback['reason']}"
            )
        return fallback
# =====================================================
# EXPORTAR CHAT PARA PDF (ESTILO MARKDOWN)
# =====================================================

INLINE_MD_PATTERN = re.compile(r"(`[^`]+`|\*\*[^*]+\*\*|\*[^*]+\*|__[^_]+__|_[^_]+_)")
LINK_MD_PATTERN = re.compile(r"\[([^\]]+)\]\(([^)]+)\)")
TABLE_SEPARATOR_RE = re.compile(r"^\s*\|?(?:\s*:?-{3,}:?\s*\|)+\s*:?-{3,}:?\s*\|?\s*$")
BULLET_LINE_RE = re.compile(r"^[-*]\s+")
ORDERED_LINE_RE = re.compile(r"^\d+\.\s+")
FENCE_RE = re.compile(r"^```")
HR_RE = re.compile(r"^\s{0,3}([-*_])(?:\s*\1){2,}\s*$")


def clean_markdown(text):

    if not text:
        return ""

    return str(text).strip()


def render_inline_markdown_pdf(text):

    if text is None:
        return ""

    raw_text = str(text)
    raw_text = LINK_MD_PATTERN.sub(r"\1 (\2)", raw_text)

    parts = []
    last = 0

    for match in INLINE_MD_PATTERN.finditer(raw_text):
        token = match.group(0)
        parts.append(escape(raw_text[last:match.start()]))

        if token.startswith("`") and token.endswith("`"):
            parts.append(f"<font name='Courier'>{escape(token[1:-1])}</font>")
        elif token.startswith("**") and token.endswith("**"):
            parts.append(f"<b>{escape(token[2:-2])}</b>")
        elif token.startswith("__") and token.endswith("__"):
            parts.append(f"<b>{escape(token[2:-2])}</b>")
        elif token.startswith("*") and token.endswith("*"):
            parts.append(f"<i>{escape(token[1:-1])}</i>")
        elif token.startswith("_") and token.endswith("_"):
            parts.append(f"<i>{escape(token[1:-1])}</i>")
        else:
            parts.append(escape(token))

        last = match.end()

    parts.append(escape(raw_text[last:]))
    return "".join(parts)


def _flush_pdf_paragraph_buffer(story, buffer, style_map):
    if not buffer:
        return

    paragraph_text = re.sub(r"\s+", " ", " ".join(buffer)).strip()
    if paragraph_text:
        story.append(Paragraph(render_inline_markdown_pdf(paragraph_text), style_map["body"]))
    buffer.clear()


def _split_markdown_table_row(line):
    row = str(line or "").strip()
    if row.startswith("|"):
        row = row[1:]
    if row.endswith("|"):
        row = row[:-1]
    return [cell.strip() for cell in row.split("|")]


def _looks_like_markdown_table(lines, idx):
    if idx + 1 >= len(lines):
        return False

    header = str(lines[idx] or "").strip()
    delimiter = str(lines[idx + 1] or "").strip()

    return "|" in header and bool(TABLE_SEPARATOR_RE.match(delimiter))


def _estimate_table_widths(rows, available_width):
    if not rows:
        return None

    col_count = max(len(row) for row in rows)
    weights = []
    for col_idx in range(col_count):
        weight = 1
        for row in rows:
            if col_idx < len(row):
                cell_len = len(re.sub(r"\s+", " ", str(row[col_idx] or "")).strip())
                weight = max(weight, min(cell_len, 40))
        weights.append(weight)

    total_weight = sum(weights) or col_count
    min_width = 42
    widths = [max(min_width, available_width * (weight / total_weight)) for weight in weights]

    if sum(widths) > available_width:
        scale = available_width / sum(widths)
        widths = [width * scale for width in widths]

    return widths


def _append_markdown_table(story, table_lines, style_map, available_width):
    if len(table_lines) < 2:
        return

    raw_rows = [_split_markdown_table_row(table_lines[0])]
    for raw_line in table_lines[2:]:
        if str(raw_line or "").strip():
            raw_rows.append(_split_markdown_table_row(raw_line))

    if not raw_rows:
        return

    col_count = max(len(row) for row in raw_rows)
    normalized_rows = [
        row + [""] * (col_count - len(row))
        for row in raw_rows
    ]

    table_data = []
    for row_idx, row in enumerate(normalized_rows):
        cell_style = style_map["table_header"] if row_idx == 0 else style_map["table_cell"]
        table_data.append([
            Paragraph(render_inline_markdown_pdf(cell or " "), cell_style)
            for cell in row
        ])

    table = Table(
        table_data,
        colWidths=_estimate_table_widths(normalized_rows, available_width),
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E9EEF6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#22324A")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.2),
        ("LEADING", (0, 0), (-1, -1), 12),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C9D3E1")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.7, colors.HexColor("#A7B8D0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F9FC")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    story.append(Spacer(1, 6))


def _append_code_block(story, code_lines, style_map, available_width):
    code_text = "\n".join(code_lines).rstrip()
    if not code_text:
        return

    code_flowable = Preformatted(code_text, style_map["code"], maxLineLength=120)
    code_box = Table(
        [[code_flowable]],
        colWidths=[available_width],
        hAlign="LEFT",
    )
    code_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F6F8")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D3DAE6")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(code_box)
    story.append(Spacer(1, 6))


def _append_list_block(story, lines, start_idx, style_map):
    idx = start_idx

    while idx < len(lines):
        stripped = str(lines[idx] or "").strip()
        if BULLET_LINE_RE.match(stripped):
            marker = "• "
            content = stripped[2:].strip()
        elif ORDERED_LINE_RE.match(stripped):
            prefix = ORDERED_LINE_RE.match(stripped).group(0)
            marker = prefix
            content = stripped[len(prefix):].strip()
        else:
            break

        idx += 1
        continuation = []
        while idx < len(lines):
            next_line = str(lines[idx] or "").rstrip()
            next_stripped = next_line.strip()
            if (
                not next_stripped
                or BULLET_LINE_RE.match(next_stripped)
                or ORDERED_LINE_RE.match(next_stripped)
                or _looks_like_markdown_table(lines, idx)
                or next_stripped.startswith("#")
                or FENCE_RE.match(next_stripped)
                or HR_RE.match(next_stripped)
            ):
                break
            continuation.append(next_stripped)
            idx += 1

        if continuation:
            content = f"{content} {' '.join(continuation)}".strip()

        story.append(Paragraph(
            f"{escape(marker)}{render_inline_markdown_pdf(content)}",
            style_map["bullet"],
        ))

    story.append(Spacer(1, 2))
    return idx


def append_markdown_block(story, text, style_map, available_width):

    if not text:
        story.append(Paragraph("_(vazio)_", style_map["body"]))
        return

    lines = str(text).replace("\r\n", "\n").replace("\r", "\n").split("\n")
    paragraph_buffer = []
    idx = 0

    while idx < len(lines):
        raw_line = str(lines[idx] or "").rstrip()
        line = raw_line.strip()

        if not line:
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            story.append(Spacer(1, 5))
            idx += 1
            continue

        if _looks_like_markdown_table(lines, idx):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            table_lines = [lines[idx], lines[idx + 1]]
            idx += 2
            while idx < len(lines):
                candidate = str(lines[idx] or "").strip()
                if not candidate or "|" not in candidate:
                    break
                table_lines.append(lines[idx])
                idx += 1
            _append_markdown_table(story, table_lines, style_map, available_width)
            continue

        if FENCE_RE.match(line):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            idx += 1
            code_lines = []
            while idx < len(lines):
                candidate = str(lines[idx] or "").rstrip()
                if FENCE_RE.match(candidate.strip()):
                    idx += 1
                    break
                code_lines.append(candidate)
                idx += 1
            _append_code_block(story, code_lines, style_map, available_width)
            continue

        if line.startswith("### "):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            story.append(Paragraph(render_inline_markdown_pdf(line[4:]), style_map["h3"]))
            idx += 1
            continue

        if line.startswith("## "):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            story.append(Paragraph(render_inline_markdown_pdf(line[3:]), style_map["h2"]))
            idx += 1
            continue

        if line.startswith("# "):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            story.append(Paragraph(render_inline_markdown_pdf(line[2:]), style_map["h1"]))
            idx += 1
            continue

        if HR_RE.match(line):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            story.append(HRFlowable(width="100%", color=colors.HexColor("#D7DEE8"), thickness=0.6))
            story.append(Spacer(1, 5))
            idx += 1
            continue

        if BULLET_LINE_RE.match(line) or ORDERED_LINE_RE.match(line):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            idx = _append_list_block(story, lines, idx, style_map)
            continue

        if line.startswith(">"):
            _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)
            quote_lines = [line.lstrip(">").strip()]
            idx += 1
            while idx < len(lines):
                candidate = str(lines[idx] or "").strip()
                if not candidate.startswith(">"):
                    break
                quote_lines.append(candidate.lstrip(">").strip())
                idx += 1
            story.append(Paragraph(
                render_inline_markdown_pdf(" ".join(quote_lines)),
                style_map["quote"],
            ))
            story.append(Spacer(1, 4))
            continue

        paragraph_buffer.append(line)
        idx += 1

    _flush_pdf_paragraph_buffer(story, paragraph_buffer, style_map)


def _build_metadata_table(item, style_map, available_width):
    rows = [
        [Paragraph("<b>Base</b>", style_map["table_key"]), Paragraph(escape(str(item.get("base", ""))), style_map["table_value"])],
        [Paragraph("<b>Modo</b>", style_map["table_key"]), Paragraph(escape(str(item.get("modo", ""))), style_map["table_value"])],
        [Paragraph("<b>Data/Hora</b>", style_map["table_key"]), Paragraph(escape(str(item.get("timestamp", ""))), style_map["table_value"])],
    ]

    metadata_table = Table(
        rows,
        colWidths=[available_width * 0.22, available_width * 0.78],
        hAlign="LEFT",
    )
    metadata_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EEF2F7")),
        ("BACKGROUND", (1, 0), (1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D8E0EA")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return metadata_table


def _draw_pdf_page_frame(canvas, doc):
    canvas.saveState()
    width, height = A4

    canvas.setStrokeColor(colors.HexColor("#D7DEE8"))
    canvas.setLineWidth(0.6)
    canvas.line(doc.leftMargin, height - 16 * mm, width - doc.rightMargin, height - 16 * mm)
    canvas.line(doc.leftMargin, 13 * mm, width - doc.rightMargin, 13 * mm)

    canvas.setFillColor(colors.HexColor("#4B5563"))
    canvas.setFont("Helvetica", 9)
    canvas.drawString(doc.leftMargin, height - 12.3 * mm, "Relatorio de Consultas RAG")
    canvas.drawRightString(width - doc.rightMargin, 9 * mm, f"Pagina {canvas.getPageNumber()}")
    canvas.restoreState()


def save_chat_pdf(chat_history, base, evidence_state=None):

    output_dir = get_user_exports_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    safe_base = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(base)).strip("_") or "rag"
    filename = f"{safe_base}_{timestamp}.pdf"
    filepath = output_dir / filename

    base_styles = getSampleStyleSheet()
    style_map = {
        "title": ParagraphStyle(
            "TitleMd",
            parent=base_styles["Title"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#1F2937"),
            spaceAfter=10,
        ),
        "h1": ParagraphStyle(
            "H1Md",
            parent=base_styles["Heading1"],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#1F2937"),
            spaceBefore=10,
            spaceAfter=6,
        ),
        "h2": ParagraphStyle(
            "H2Md",
            parent=base_styles["Heading2"],
            fontSize=12.5,
            leading=16,
            textColor=colors.HexColor("#243B53"),
            spaceBefore=7,
            spaceAfter=4,
        ),
        "h3": ParagraphStyle(
            "H3Md",
            parent=base_styles["Heading3"],
            fontSize=11.5,
            leading=14,
            textColor=colors.HexColor("#334E68"),
            spaceBefore=4,
            spaceAfter=3,
        ),
        "body": ParagraphStyle(
            "BodyMd",
            parent=base_styles["BodyText"],
            fontSize=10.5,
            leading=14,
            textColor=colors.HexColor("#1F2937"),
            alignment=4,
            spaceAfter=4,
        ),
        "bullet": ParagraphStyle(
            "BulletMd",
            parent=base_styles["BodyText"],
            fontSize=10.5,
            leading=14,
            leftIndent=14,
            firstLineIndent=-8,
            textColor=colors.HexColor("#1F2937"),
            spaceAfter=2,
        ),
        "meta": ParagraphStyle(
            "MetaMd",
            parent=base_styles["BodyText"],
            fontSize=9,
            textColor=colors.HexColor("#555555"),
            spaceAfter=3,
        ),
        "quote": ParagraphStyle(
            "QuoteMd",
            parent=base_styles["BodyText"],
            fontSize=10.2,
            leading=14,
            leftIndent=12,
            rightIndent=8,
            borderPadding=6,
            borderColor=colors.HexColor("#CBD5E1"),
            borderWidth=0.6,
            borderLeft=True,
            textColor=colors.HexColor("#334155"),
            backColor=colors.HexColor("#F8FAFC"),
            spaceAfter=5,
        ),
        "code": ParagraphStyle(
            "CodeMd",
            parent=base_styles["Code"],
            fontName="Courier",
            fontSize=8.7,
            leading=11,
            textColor=colors.HexColor("#1F2937"),
        ),
        "table_header": ParagraphStyle(
            "TableHeaderMd",
            parent=base_styles["BodyText"],
            fontSize=9.3,
            leading=11.5,
            textColor=colors.HexColor("#22324A"),
        ),
        "table_cell": ParagraphStyle(
            "TableCellMd",
            parent=base_styles["BodyText"],
            fontSize=9.1,
            leading=11.5,
            textColor=colors.HexColor("#1F2937"),
        ),
        "table_key": ParagraphStyle(
            "TableKeyMd",
            parent=base_styles["BodyText"],
            fontSize=9.2,
            leading=12,
            textColor=colors.HexColor("#334155"),
        ),
        "table_value": ParagraphStyle(
            "TableValueMd",
            parent=base_styles["BodyText"],
            fontSize=9.2,
            leading=12,
            textColor=colors.HexColor("#1F2937"),
        ),
    }

    story = []

    story.append(Paragraph("Relatorio de Consultas RAG", style_map["title"]))
    story.append(Paragraph(f"Base exportada: {escape(str(base))}", style_map["meta"]))
    story.append(Paragraph(f"Gerado em: {escape(generated_at)}", style_map["meta"]))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#D7DEE8"), thickness=0.8))
    story.append(Spacer(1, 10))

    evidence_state = evidence_state or {}
    content_width = A4[0] - (20 * mm) - (18 * mm)

    for idx, item in enumerate(chat_history, start=1):
        question = clean_markdown(item.get("question", ""))
        answer = clean_markdown(item.get("answer", ""))
        evidence = clean_markdown(item.get("evidence", ""))
        include_evidence = bool(evidence_state.get(idx - 1, False)) and bool(evidence)

        story.append(Paragraph(f"Consulta {idx}", style_map["h1"]))
        story.append(Spacer(1, 1))

        story.append(Paragraph("Pergunta", style_map["h2"]))
        append_markdown_block(story, question, style_map, content_width)

        story.append(Paragraph("Resposta", style_map["h2"]))
        append_markdown_block(story, answer, style_map, content_width)

        story.append(Paragraph("Metadados", style_map["h2"]))
        story.append(_build_metadata_table(item, style_map, content_width))
        story.append(Spacer(1, 5))

        timings = item.get("timings") or []
        if timings:
            story.append(Paragraph("Tempos", style_map["h2"]))
            for t in timings:
                story.append(Paragraph(f"• {escape(clean_markdown(t))}", style_map["bullet"]))
            story.append(Spacer(1, 4))

        if include_evidence:
            story.append(Paragraph("Evidencias", style_map["h2"]))
            append_markdown_block(story, evidence, style_map, content_width)

        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", color=colors.HexColor("#E2E8F0"), thickness=0.7))
        story.append(Spacer(1, 10))

    doc = SimpleDocTemplate(
        str(filepath),
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=18 * mm,
        topMargin=22 * mm,
        bottomMargin=18 * mm,
        title=f"Relatorio de Consultas RAG - {base}",
        author="RAG-CGENEDIV1",
    )
    doc.build(
        story,
        onFirstPage=_draw_pdf_page_frame,
        onLaterPages=_draw_pdf_page_frame,
    )

    return filepath


def render_app_header():

    image_candidates = [
        BASE_DIR / "assets" / "rag_header.png",
        BASE_DIR / "assets" / "rag_divi.png",
        BASE_DIR / "assets" / "rag_divi.jpg",
        BASE_DIR / "assets" / "rag_cgu.png",
        BASE_DIR / "rag_divi.png",
        BASE_DIR / "rag_cgu.png",
    ]

    banner_path = next((p for p in image_candidates if p.exists()), None)

    if banner_path is not None:
        if PIL_AVAILABLE:
            try:
                with Image.open(banner_path) as img:
                    img_rgb = img.convert("RGB")
                    bg_color = img_rgb.getpixel((0, 0))
                    bg = Image.new("RGB", img_rgb.size, bg_color)
                    diff = ImageChops.difference(img_rgb, bg)
                    bbox = diff.getbbox()

                    if bbox:
                        img_rgb = img_rgb.crop(bbox)

                    target_width = max(504, int(img_rgb.width * 0.72))
                    st.image(img_rgb, width=target_width)
            except Exception:
                st.image(str(banner_path), width=912)
        else:
            st.image(str(banner_path), width=912)
        return

    st.title("\U0001F6E2 RAG-CGENEDIV1 \U0001F525")
# =====================================================
# CONFIGURAO DA PÃGINA
# =====================================================

st.set_page_config(
    layout="wide",
    page_title="RAG-CGENEDIV1 \U0001F525",
    page_icon="\U0001F6E2"
)

render_app_header()


# =====================================================
# SESSION STATE
# =====================================================

defaults = {
    "filtered_docs": None,
    "filtered_docs_by_base": None,
    "lexical_base_docs": None,
    "lexical_base_docs_by_base": None,
    "filtro_ativo": False,
    "toggle_filtro": False,
    "termos_busca": "",
    "clear_termos_busca": False,
    "clear_modo_busca": False,
    "content_filters": [],
    "content_filters_version": 0,
    "clear_termos_busca_2": False,
    "termos_busca_2": "",  # NOVA VARIÃVEL
    "modo_busca": "Nome do documento",
    "doc_forcado": None,
    "modo_consulta": "Automatico (Router Inteligente)",
    "sqlite_db_choice": "",
    "last_export_xlsx": None,
    "last_export_pdf": None,
    "last_cited_export_xlsx": None,
    "last_base": None,
    "show_api_editor": False,
    "show_base_root_editor": False,
    "show_base_root_save_form": False,
    "custom_base_root": "",
    "base_root_choice": "",
    "base_root_choice_pending": None,
    "base_root_save_name": "",
    "base_root_make_default": False,
    "base_root_feedback": None,
    "base_root_error": None,
    "multi_base_enabled": False,
    "multi_selected_bases": [],
    "temperature": 0.2,
    "final_llm_model": FINAL_LLM_MODELS[0],
    "prompt_template_choice": "Juridico",
    "prompt_template_choice_pending": None,
    "prompt_save_name": "",
    "prompt_save_feedback": None,
    "memory_enabled": True,
    "memory_short_enabled": True,
    "memory_persistent_enabled": True,
    "memory_auto_summarize": True,
    "memory_short_turns": 4,
    "memory_retrieval_limit": 5,
    "memory_clear_feedback": None,
}

RAG_UI_STATE_KEYS = [
    "top_k_retrieval",
    "top_k_final",
    "alpha_semantic",
    "beta_lexical",
    "max_context_chars",
    "adaptive_tuning_enabled",
    "max_query_variants",
    "hybrid_profiles_per_query",
    "max_hybrid_searches",
    "hybrid_time_budget_s",
    "min_hybrid_calls_before_early_stop",
    "early_stop_min_unique_chunks",
    "early_stop_no_gain_patience",
    "prompt_custom",
    "final_llm_model",
    "prompt_template_choice",
    "prompt_save_name",
    "temperature",
    "modo_consulta",
    "doc_forcado",
]

for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

if (
    not st.session_state.get("custom_base_root")
    and not st.session_state.get("base_root_choice")
):
    default_saved_base_root = get_default_saved_base_root()
    if default_saved_base_root:
        st.session_state.custom_base_root = default_saved_base_root

if st.session_state.get("final_llm_model") not in FINAL_LLM_MODELS:
    st.session_state.final_llm_model = FINAL_LLM_MODELS[0]

apply_session_base_root()

# =====================================================
# BASES
# =====================================================

st.sidebar.markdown("---")

if st.sidebar.button("\U0001F512 Inserir/Alterar API"):
    st.session_state.show_api_editor = not st.session_state.show_api_editor

if st.session_state.show_api_editor:
    with st.sidebar.form("api_key_form"):
        api_input = st.text_input(
            "LIA_API_KEY",
            value=get_current_api_key(),
            type="password",
            help="A chave sera salva no .env do usuario."
        )
        save_api = st.form_submit_button("Salvar API")

    if save_api:
        api_value = api_input.strip()

        if not api_value:
            st.sidebar.warning("Informe uma chave API valida.")
        else:
            try:
                upsert_env_var(USER_ENV, "LIA_API_KEY", api_value)
                os.environ["LIA_API_KEY"] = api_value
                st.sidebar.success(f"API salva em: {USER_ENV}")
            except Exception as e:
                st.sidebar.error(f"Erro ao salvar API: {e}")

active_root = get_shared_base_root()
st.sidebar.caption(f"Raiz base_rag ativa: `{active_root}`")

base_root_feedback = st.session_state.pop("base_root_feedback", None)
if base_root_feedback:
    st.sidebar.success(base_root_feedback)
base_root_error = st.session_state.pop("base_root_error", None)
if base_root_error:
    st.sidebar.error(base_root_error)

saved_base_payload = load_saved_base_roots()
saved_base_roots = saved_base_payload["roots"]
saved_base_options = ["Padrão"] + list(saved_base_roots.keys())
default_saved_name = saved_base_payload.get("default", "")
base_root_choice_pending = st.session_state.pop("base_root_choice_pending", None)
if base_root_choice_pending:
    st.session_state.base_root_choice = base_root_choice_pending

current_choice = st.session_state.get("base_root_choice") or default_saved_name or "Padrão"
if current_choice not in saved_base_options:
    current_choice = default_saved_name if default_saved_name in saved_base_options else "Padrão"

selected_saved_root = st.sidebar.selectbox(
    "Caminho da base",
    saved_base_options,
    index=saved_base_options.index(current_choice),
    key="base_root_choice",
    format_func=lambda name: f"{name} (padrão do usuário)" if name == default_saved_name else name,
    on_change=apply_base_root_choice,
)

if selected_saved_root == "Padrão":
    selected_base_root_path = DEFAULT_SHARED_BASE_ROOT
    st.sidebar.caption(f"`{selected_base_root_path}`")
else:
    selected_base_root_path = saved_base_roots[selected_saved_root]
    st.sidebar.caption(f"`{selected_base_root_path}`")

if st.sidebar.button("Salvar novo caminho da base", use_container_width=True):
    st.session_state.show_base_root_save_form = not st.session_state.show_base_root_save_form
    if not st.session_state.custom_base_root:
        st.session_state.custom_base_root = str(active_root)

if st.session_state.show_base_root_save_form:
    with st.sidebar.form("base_root_form"):
        base_root_input = st.text_input(
            "Novo caminho da base_rag",
            value=st.session_state.get("custom_base_root") or str(active_root),
            help=(
                "Informe a pasta raiz base_rag. Ela deve conter as subpastas "
                "`documentos` e `data`."
            ),
        )
        base_root_save_name = st.text_input(
            "Salvar este caminho como",
            value=st.session_state.get("base_root_save_name", ""),
            placeholder="Ex: Rede DIV1, Base local, Homologacao",
        )
        base_root_make_default = st.checkbox(
            "Salvar como padrão do usuário",
            value=bool(st.session_state.get("base_root_make_default", False)),
            help="Quando marcado, este caminho sera selecionado automaticamente ao abrir o RAG.",
        )
        save_named_base_root = st.form_submit_button("Salvar")

    if save_named_base_root:
        root_path, error = validate_base_root_path(base_root_input)
        if error:
            st.sidebar.error(error)
        else:
            try:
                saved_name, saved_path = save_base_root_entry(
                    base_root_save_name,
                    str(root_path),
                    make_default=base_root_make_default,
                )
                st.session_state.custom_base_root = str(root_path)
                st.session_state.base_root_choice_pending = saved_name
                st.session_state.base_root_save_name = ""
                st.session_state.base_root_make_default = False
                st.session_state.show_base_root_save_form = False
                set_shared_base_root(root_path)
                clear_runtime_caches()
                st.session_state.base_root_feedback = (
                    f"Caminho '{saved_name}' salvo em: {saved_path}"
                )
                st.rerun()
            except ValueError as exc:
                st.sidebar.warning(str(exc))
            except Exception as exc:
                st.sidebar.error(f"Erro ao salvar caminho: {exc}")

st.sidebar.toggle(
    "🔀 Modo multibase",
    key="multi_base_enabled",
    help="Quando ligado, a pergunta sera consultada em mais de uma base e as respostas serao agrupadas.",
)

bases = get_available_bases()

if not bases:
    st.warning("Nenhuma base encontrada.")
    st.stop()

if st.session_state.multi_base_enabled:
    default_multi = st.session_state.get("multi_selected_bases") or bases[: min(2, len(bases))]
    default_multi = [item for item in default_multi if item in bases]
    if not default_multi:
        default_multi = bases[: min(2, len(bases))]
    selected_bases = st.sidebar.multiselect(
        "\U0001F4C2 Selecionar Bases",
        bases,
        default=default_multi,
        key="multi_selected_bases",
    )
    if not selected_bases:
        selected_bases = default_multi or bases[:1]
        st.sidebar.warning("Selecione ao menos uma base para o modo multibase.")
    base = selected_bases[0]
    st.sidebar.caption(f"Multibase ativo: {len(selected_bases)} bases selecionadas.")
else:
    selected_bases = []
    base = st.sidebar.selectbox("\U0001F4C2 Selecionar Base", bases)
# =====================================================
# CARREGAR ÃNDICE LEXICAL
# =====================================================

@st.cache_resource(show_spinner=False)
def load_index(base):

    try:
        return load_positional_index(base)
    except FileNotFoundError:
        return None


@st.cache_data(show_spinner=False)
def load_doc_text_lookup(base):

    chunks = load_valid_jsonl(get_chunks_path(base))
    doc_text = {}

    for chunk in chunks:

        doc = chunk.get("doc")
        text = chunk.get("text", "")

        if not doc:
            continue

        if doc not in doc_text:
            doc_text[doc] = []

        doc_text[doc].append(text)

    return {
        doc: "\n".join(parts)
        for doc, parts in doc_text.items()
    }


index = load_index(base)

if index is None:
    st.warning(
        "Esta base ainda nao foi indexada.\n\n"
        "Execute a indexacao fora do app com: `python index.py --base <nome_da_base>`."
    )

# =====================================================
# RESET AO TROCAR BASE
# =====================================================

if st.session_state.last_base != base:

    # preservar parÃ¢metros do RAG e prompt
    rag_backup = {
        key: st.session_state.get(key)
        for key in RAG_UI_STATE_KEYS
        if key in st.session_state
    }

    # reset apenas estado dependente da base
    st.session_state.filtered_docs = None
    st.session_state.filtered_docs_by_base = None
    st.session_state.lexical_base_docs = None
    st.session_state.lexical_base_docs_by_base = None
    st.session_state.filtro_ativo = False
    st.session_state.toggle_filtro = False
    st.session_state.termos_busca = ""
    st.session_state.termos_busca_2 = ""
    st.session_state.content_filters = []
    st.session_state.content_filters_version += 1
    st.session_state.modo_busca = "Nome do documento"
    st.session_state.doc_forcado = None
    st.session_state.sqlite_db_choice = ""
    st.session_state.last_export_xlsx = None
    st.session_state.last_export_pdf = None
    st.session_state.last_cited_export_xlsx = None
    st.session_state.last_cited_export_xlsx = None

    # modo de consulta volta ao padrÃ£o
    st.session_state.modo_consulta = "Automatico (Router Inteligente)"

    # restaurar parÃ¢metros RAG
    for k, v in rag_backup.items():
        if v is not None:
            st.session_state[k] = v

    st.session_state.last_base = base


# =====================================================
# CAMINHO DOCUMENTOS
# =====================================================

documentos_dir = get_documentos_root() / base
documentos_dir.mkdir(parents=True, exist_ok=True)


# =====================================================
# ABRIR PASTA
# =====================================================

st.sidebar.markdown("---")

if st.sidebar.button("\U0001F4C2 Abrir Pasta de Documentos"):

    if sys.platform == "win32":
        os.startfile(documentos_dir)
    else:
        subprocess.Popen(["xdg-open", str(documentos_dir)])


# =====================================================
# SALVAR CONSULTAS EM PDF
# =====================================================

st.sidebar.caption("Evidencias sao exportadas apenas quando estiverem visiveis no chat.")

if st.sidebar.button("\U0001F4BE Salvar consultas em PDF"):

    if st.session_state.get("chat_history"):

        try:

            pdf_path = save_chat_pdf(
                st.session_state.chat_history,
                get_active_base_label(),
                st.session_state.get("evidence_state", {})
            )
            st.session_state.last_export_pdf = str(pdf_path)

            st.sidebar.success(f"PDF salvo em:\n{pdf_path}")

        except Exception as e:

            st.sidebar.error(f"Erro ao gerar PDF: {e}")

    else:

        st.sidebar.warning("Nenhuma consulta para salvar.")

pdf_export_path = st.session_state.get("last_export_pdf")

if pdf_export_path:
    pdf_export_path = Path(pdf_export_path)
    if pdf_export_path.exists():
        st.sidebar.caption(f"Arquivo: {pdf_export_path}")
        if st.sidebar.button("\U0001F4C2 Abrir pasta da exportacao PDF"):
            try:
                export_dir = pdf_export_path.parent
                if sys.platform == "win32":
                    os.startfile(str(export_dir))
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", str(export_dir)])
                else:
                    subprocess.Popen(["xdg-open", str(export_dir)])
            except Exception as e:
                st.sidebar.error(f"Erro ao abrir pasta da exportacao: {e}")


if st.sidebar.button(
    "\U0001F4BE Salvar documentos citados em Excel",
    disabled=not OPENPYXL_AVAILABLE
):

    if st.session_state.get("chat_history"):

        try:

            cited_item = st.session_state.chat_history[-1]
            response_index = len(st.session_state.chat_history) - 1
            current_pdfs = list(documentos_dir.glob("*.pdf"))
            available_docs_set, available_docs_by_key = build_doc_name_maps(current_pdfs)

            cited_export_path = export_cited_docs_excel(
                cited_item,
                response_index,
                base,
                documentos_dir,
                available_docs_set,
                available_docs_by_key,
            )
            st.session_state.last_cited_export_xlsx = str(cited_export_path)

            st.sidebar.success(
                f"Excel de documentos citados salvo em:\n{cited_export_path}"
            )

        except Exception as e:

            st.sidebar.error(f"Erro ao gerar Excel de documentos citados: {e}")

    else:

        st.sidebar.warning("Nenhuma resposta do RAG para exportar.")

cited_export_path = st.session_state.get("last_cited_export_xlsx")

if cited_export_path:
    cited_export_path = Path(cited_export_path)
    if cited_export_path.exists():
        st.sidebar.caption(f"Documentos citados: {cited_export_path}")
        if st.sidebar.button("\U0001F4C2 Abrir pasta dos documentos citados"):
            try:
                export_dir = cited_export_path.parent
                if sys.platform == "win32":
                    os.startfile(str(export_dir))
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", str(export_dir)])
                else:
                    subprocess.Popen(["xdg-open", str(export_dir)])
            except Exception as e:
                st.sidebar.error(f"Erro ao abrir pasta da exportacao: {e}")


# =====================================================
# ESTATÃSTICAS
# =====================================================

chunks_path = get_chunks_path(base)

pdfs = list(documentos_dir.glob("*.pdf"))
num_docs = len(pdfs)
scope_num_docs = get_total_docs_for_active_scope()

def get_chunk_count(path):
    if path.exists():
        return len(load_valid_jsonl(path))
    return 0

num_chunks = get_chunk_count(chunks_path)

st.sidebar.markdown("---")
st.sidebar.markdown("### \U0001F4CA Estatisticas")

st.sidebar.write(f"\U0001F4C4 Documentos: {scope_num_docs}")
st.sidebar.write(f"\U0001F9E9 Chunks: {num_chunks}")

# =====================================================
# MODO DE CONSULTA
# =====================================================

st.sidebar.markdown("---")
st.sidebar.markdown("### \U0001F50E Modo de Consulta")

if st.session_state.get("filtered_docs") is not None:
    doc_list = st.session_state.filtered_docs
else:
    doc_list = [f.name for f in pdfs]

modo_consulta = st.sidebar.radio(
    "Escolha o modo:",
    [
        "Automatico (Router Inteligente)",
        "Forcar Busca Global",
        "Forcar Documento Especifico",
        "Somente SQLite da base"
    ],
    key="modo_consulta"
)

doc_forcado = None
sqlite_files = get_sqlite_files(base)
selected_sqlite_path = None

if sqlite_files:
    sqlite_options = [str(path) for path in sqlite_files]
    current_sqlite = st.session_state.get("sqlite_db_choice")
    if current_sqlite not in sqlite_options:
        st.session_state.sqlite_db_choice = sqlite_options[0]
    selected_sqlite_path = Path(
        st.sidebar.selectbox(
            "SQLite da base",
            sqlite_options,
            key="sqlite_db_choice",
            format_func=lambda value: Path(value).name,
            help="Usado no modo 'Somente SQLite da base' e nos comandos @sqlite/@sql.",
        )
    )
else:
    st.session_state.sqlite_db_choice = ""
    if modo_consulta == "Somente SQLite da base":
        st.sidebar.warning("Nenhum SQLite encontrado em base_rag/data desta base. Use o app4 para criar.")

if modo_consulta == "Forcar Documento Especifico":

    if doc_list:

        doc_forcado = st.sidebar.selectbox(
            "Selecionar documento:",
            doc_list,
            key=f"doc_select_{base}_{len(doc_list)}"
        )

    else:

        st.sidebar.warning(
            "Nenhum documento disponÃ­vel na base filtrada atual."
        )


# =====================================================
# TEMPERATURA
# =====================================================

st.sidebar.markdown("---")

temperature = st.sidebar.slider(
    "\U0001F321 Temperatura",
    0.0,
    1.0,
    key="temperature",
)

final_llm_model = st.sidebar.selectbox(
    "Modelo LLM da resposta final",
    FINAL_LLM_MODELS,
    key="final_llm_model",
    help="Modelo usado apenas na etapa final de analise/resposta.",
)


# =====================================================
# FILTRO LEXICAL
# =====================================================

st.sidebar.markdown("---")
st.sidebar.markdown("### \U0001F50E Delimitar Base Lexicalmente (Filtro Booleano)")

if st.session_state.get("clear_termos_busca"):
    st.session_state.termos_busca = ""
    st.session_state.clear_termos_busca = False

filtro_on = st.sidebar.toggle(
    "Ativar filtro lexical",
    key="toggle_filtro"
)

termos = st.sidebar.text_input(
    "Palavras ou termos",
    key="termos_busca",
    placeholder="Ex: PDI custo abandono coral"
)

if st.session_state.get("clear_modo_busca"):
    st.session_state.modo_busca = "Nome do documento"
    st.session_state.clear_modo_busca = False

modo_busca = st.sidebar.radio(
    "Buscar em:",
    [
        "Nome do documento",
        "Nome + conte\u00fado"
    ],
    key="modo_busca"
)


# =====================================================
# RESET AO DESATIVAR FILTRO
# =====================================================

if (
    not filtro_on and (
        st.session_state.get("filtro_ativo")
        or st.session_state.get("filtered_docs") is not None
        or st.session_state.get("filtered_docs_by_base") is not None
        or st.session_state.get("lexical_base_docs") is not None
        or st.session_state.get("lexical_base_docs_by_base") is not None
    )
):

    rag_backup = {
        "top_k_retrieval": st.session_state.get("top_k_retrieval"),
        "top_k_final": st.session_state.get("top_k_final"),
        "alpha_semantic": st.session_state.get("alpha_semantic"),
        "beta_lexical": st.session_state.get("beta_lexical"),
        "max_context_chars": st.session_state.get("max_context_chars"),
        "prompt_custom": st.session_state.get("prompt_custom")
    }

    st.session_state.filtered_docs = None
    st.session_state.filtered_docs_by_base = None
    st.session_state.lexical_base_docs = None
    st.session_state.lexical_base_docs_by_base = None
    st.session_state.filtro_ativo = False
    st.session_state.clear_termos_busca = True
    st.session_state.termos_busca_2 = ""
    st.session_state.content_filters = []
    st.session_state.content_filters_version += 1
    st.session_state.clear_modo_busca = True
    st.session_state.doc_forcado = None
    st.session_state.last_export_xlsx = None
    st.session_state.last_export_pdf = None

    for k, v in rag_backup.items():
        if v is not None:
            st.session_state[k] = v

    st.rerun()


# =====================================================
# APLICAR FILTRO
# =====================================================

if filtro_on:

    if st.sidebar.button("Aplicar filtro lexical"):

        # preservar parÃ¢metros do RAG
        rag_backup = {
            "top_k_retrieval": st.session_state.get("top_k_retrieval"),
            "top_k_final": st.session_state.get("top_k_final"),
            "alpha_semantic": st.session_state.get("alpha_semantic"),
            "beta_lexical": st.session_state.get("beta_lexical"),
            "max_context_chars": st.session_state.get("max_context_chars"),
            "prompt_custom": st.session_state.get("prompt_custom")
        }

        st.session_state.filtro_ativo = True

        # =====================================================
        # BUSCA NO ÃNDICE LEXICAL (conteÃºdo)
        # =====================================================

        terms, operator, exclude = parse_boolean_query(termos)

        filter_bases = get_filter_base_names()
        lexical_by_base = {}

        for filter_base in filter_bases:
            base_index = load_index(filter_base)

            if base_index is None:
                lexical_by_base[filter_base] = []
                continue

            docs_conteudo = docs_from_lexical_query_parts(
                base_index,
                terms,
                operator,
                exclude,
            )

            # =====================================================
            # BUSCA NO NOME DO DOCUMENTO
            # =====================================================

            docs_nome_match = []
            base_documentos_dir = get_documentos_root() / filter_base

            for pdf in base_documentos_dir.glob("*.pdf"):

                if doc_name_matches_lexical_query(
                    pdf.name,
                    termos,
                    terms,
                    operator,
                    exclude,
                ):
                    docs_nome_match.append(pdf.name)

            # =====================================================
            # COMBINACAO CONFORME MODO DE BUSCA
            # =====================================================

            if modo_busca == "Nome do documento":

                filtrados_base = docs_nome_match

            else:

                filtrados_base = list(
                    set(docs_nome_match) | docs_conteudo
                )

            lexical_by_base[filter_base] = list(filtrados_base)

        filtrados = lexical_by_base.get(base, [])

        # =====================================================
        # SALVAR RESULTADO
        # =====================================================

        st.session_state.lexical_base_docs = list(filtrados)
        st.session_state.lexical_base_docs_by_base = lexical_by_base
        st.session_state.content_filters = []
        st.session_state.content_filters_version += 1
        st.session_state.termos_busca_2 = ""
        recompute_content_refinement()

        # restaurar parÃ¢metros do RAG
        for k, v in rag_backup.items():
            if v is not None:
                st.session_state[k] = v

        st.sidebar.success(
            f"{sum(len(docs) for docs in lexical_by_base.values())} documentos encontrados."
        )
# =====================================================
# BASE FILTRADA
# =====================================================

if st.session_state.get("filtered_docs") is not None:

    st.sidebar.markdown("### \U0001F4CA Base delimitada")
    st.sidebar.write(f"\U0001F4C4 Base total: {scope_num_docs}")
    st.sidebar.write(f"\U0001F50E Base filtrada: {get_filtered_docs_total()}")


# =====================================================
# SEGUNDA CAMADA DE FILTRO
# =====================================================

if st.session_state.get("lexical_base_docs") is not None:

    st.sidebar.markdown("---")
    st.sidebar.markdown("### \U0001F50E Refinar base pelo conte\u00fado")

    if st.session_state.get("clear_termos_busca_2"):
        st.session_state.termos_busca_2 = ""
        st.session_state.clear_termos_busca_2 = False

    termos_2 = st.sidebar.text_input(
        "Palavras no conte\u00fado",
        key="termos_busca_2",
        placeholder='Ex: "plano de abandono" AND ibama'
    )

    if st.sidebar.button("Aplicar refinamento de conte\u00fado"):

        if termos_2.strip():

            try:

                docs_antes_total = get_filtered_docs_total()
                if docs_antes_total == 0:
                    lexical_by_base = st.session_state.get("lexical_base_docs_by_base")
                    if st.session_state.get("multi_base_enabled") and lexical_by_base is not None:
                        docs_antes_total = sum(
                            len(lexical_by_base.get(base_name, []))
                            for base_name in get_filter_base_names()
                        )
                    else:
                        docs_antes_total = len(st.session_state.get("lexical_base_docs") or [])

                docs_by_base = {}

                for filter_base in get_filter_base_names():
                    base_index = load_index(filter_base)
                    if base_index is None:
                        docs_by_base[filter_base] = []
                        continue

                    terms_2, operator_2, exclude_2 = parse_boolean_query(
                        termos_2.strip()
                    )
                    docs_by_base[filter_base] = sorted(
                        docs_from_lexical_query_parts(
                            base_index,
                            terms_2,
                            operator_2,
                            exclude_2,
                        )
                    )

                docs_conteudo = docs_by_base.get(base, [])

                filtros = list(st.session_state.get("content_filters", []))
                term_norm = termos_2.strip().lower()
                filtro_existente = None

                for i, filtro in enumerate(filtros):
                    if filtro.get("term", "").strip().lower() == term_norm:
                        filtro_existente = i
                        break

                if filtro_existente is not None:
                    filtros[filtro_existente]["docs"] = docs_conteudo
                    filtros[filtro_existente]["docs_by_base"] = docs_by_base
                    filtros[filtro_existente]["active"] = True
                else:
                    filtros.append({
                        "term": termos_2.strip(),
                        "docs": docs_conteudo,
                        "docs_by_base": docs_by_base,
                        "active": True
                    })

                st.session_state.content_filters = filtros
                st.session_state.content_filters_version += 1
                recompute_content_refinement()

                docs_depois_total = get_filtered_docs_total()
                docs_removidos = docs_antes_total - docs_depois_total

                if docs_removidos > 0:
                    st.sidebar.success(
                        f"Base refinada: {docs_antes_total} -> "
                        f"{docs_depois_total} documentos "
                        f"({docs_removidos} removidos)."
                    )
                elif is_negative_only_filter(termos_2.strip()):
                    st.sidebar.info(
                        f"Filtro aplicado, mas nenhum documento foi removido: "
                        f"os {docs_antes_total} documentos atuais nao contem "
                        "o termo excluido."
                    )
                else:
                    st.sidebar.warning(
                        f"Filtro aplicado sem reduzir a base: "
                        f"{docs_antes_total} -> {docs_depois_total} documentos."
                    )

            except Exception as e:

                st.sidebar.error(f"Erro no refinamento: {e}")

    if st.session_state.get("content_filters"):

        st.sidebar.markdown("#### Palavras/termos usados")

        filtros_alterados = False
        filtros = st.session_state.content_filters
        version = st.session_state.get("content_filters_version", 0)
        docs_atuais = st.session_state.get("filtered_docs") or []
        doc_text_lookup = None
        if not st.session_state.get("multi_base_enabled"):
            doc_text_lookup = load_doc_text_lookup(base)

        for i, filtro in enumerate(filtros):

            if st.session_state.get("multi_base_enabled"):
                docs_by_base = st.session_state.get("filtered_docs_by_base") or {}
                total_ocorrencias = sum(
                    get_query_total_occurrences(
                        filtro.get("term", ""),
                        docs_by_base.get(base_name, []),
                        load_doc_text_lookup(base_name),
                    )
                    for base_name in get_filter_base_names()
                )
            else:
                total_ocorrencias = get_query_total_occurrences(
                    filtro.get("term", ""),
                    docs_atuais,
                    doc_text_lookup
                )

            label_filtro = describe_lexical_filter(filtro.get("term", ""))
            ocorrencias_label = (
                "ocorrencias restantes"
                if is_negative_only_filter(filtro.get("term", ""))
                else "ocorrencias"
            )

            ativo = st.sidebar.toggle(
                f"{label_filtro} ({total_ocorrencias} {ocorrencias_label})",
                value=filtro.get("active", True),
                key=f"content_filter_active_{base}_{version}_{i}"
            )

            if ativo != filtro.get("active", True):
                st.session_state.content_filters[i]["active"] = ativo
                filtros_alterados = True

        if filtros_alterados:
            recompute_content_refinement()
            st.rerun()

        if st.sidebar.button("Limpar refinamentos de conte\u00fado"):
            st.session_state.content_filters = []
            st.session_state.content_filters_version += 1
            st.session_state.clear_termos_busca_2 = True
            recompute_content_refinement()
            st.sidebar.success("Refinamentos removidos.")
            st.rerun()

    st.sidebar.markdown("#### Exportar base filtrada")

    if not OPENPYXL_AVAILABLE:
        st.sidebar.warning(
            "Pacote openpyxl nao instalado. "
            "Instale para habilitar a exportacao em Excel."
        )

    if st.sidebar.button(
        "Exportar base filtrada (Excel)",
        disabled=not OPENPYXL_AVAILABLE
    ):

        try:

            termos_export = collect_active_terms()

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Base Filtrada"

            headers = ["base", "documento"] + termos_export + ["ocorrencias_total"]
            sheet.append(headers)

            if st.session_state.get("multi_base_enabled"):
                docs_by_base = st.session_state.get("filtered_docs_by_base") or {}
                export_items = [
                    (base_name, doc)
                    for base_name in get_filter_base_names()
                    for doc in docs_by_base.get(base_name, [])
                ]
            else:
                export_items = [
                    (base, doc)
                    for doc in (st.session_state.get("filtered_docs") or [])
                ]

            doc_text_lookup_cache = {}
            doc_maps_cache = {}

            for item_base, doc in export_items:

                if item_base not in doc_text_lookup_cache:
                    doc_text_lookup_cache[item_base] = load_doc_text_lookup(item_base)

                if item_base not in doc_maps_cache:
                    item_dir = get_documentos_root() / item_base
                    item_pdfs = list(item_dir.glob("*.pdf"))
                    doc_maps_cache[item_base] = build_doc_name_maps(item_pdfs)

                doc_text_lookup = doc_text_lookup_cache[item_base]
                available_docs_set, available_docs_by_key = doc_maps_cache[item_base]

                resolved_doc = resolve_doc_name(
                    doc,
                    available_docs_set,
                    available_docs_by_key
                )

                doc_path = get_documentos_root() / item_base / resolved_doc
                doc_text = (
                    doc_text_lookup.get(doc)
                    or doc_text_lookup.get(resolved_doc)
                    or doc_text_lookup.get(fix_mojibake_text(doc), "")
                )

                row = [item_base, resolved_doc]
                total = 0

                for term in termos_export:

                    ocorrencias = count_term_occurrences(doc_text, term)
                    row.append(ocorrencias)
                    total += ocorrencias

                row.append(total)
                sheet.append(row)

                doc_cell = sheet.cell(row=sheet.max_row, column=1)
                doc_cell.value = item_base

                doc_cell = sheet.cell(row=sheet.max_row, column=2)
                doc_cell.value = resolved_doc
                doc_cell.hyperlink = str(doc_path)
                doc_cell.style = "Hyperlink"

            export_dir = get_user_exports_dir()
            export_dir.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_base = "multibase" if st.session_state.get("multi_base_enabled") else base
            export_path = export_dir / f"{export_base}_base_filtrada_{timestamp}.xlsx"

            workbook.save(export_path)

            st.session_state.last_export_xlsx = str(export_path)

            st.sidebar.success(f"Excel exportado: {export_path.name}")

        except Exception as e:

            st.sidebar.error(f"Erro ao exportar Excel: {e}")

    export_path_str = st.session_state.get("last_export_xlsx")

    if export_path_str:

        export_path = Path(export_path_str)

        if export_path.exists():

            st.sidebar.caption(f"Arquivo: {export_path}")

            if st.sidebar.button("Abrir pasta da exportacao"):

                try:

                    export_dir = export_path.parent

                    if sys.platform == "win32":
                        os.startfile(str(export_dir))
                    elif sys.platform == "darwin":
                        subprocess.Popen(["open", str(export_dir)])
                    else:
                        subprocess.Popen(["xdg-open", str(export_dir)])

                except Exception as e:

                    st.sidebar.error(f"Erro ao abrir exportacao: {e}")

    # =====================================================
    # MOSTRAR BASE DELIMITADA ATUALIZADA
    # =====================================================

    st.sidebar.markdown("### \U0001F4CA Base delimitada")
    st.sidebar.write(f"\U0001F4C4 Base total: {scope_num_docs}")
    st.sidebar.write(f"\U0001F50E Base filtrada: {get_filtered_docs_total()}")

# =====================================================
# DOCUMENTOS DA BASE
# =====================================================

st.sidebar.markdown("---")
st.sidebar.markdown("### \U0001F4C2 Documentos da Base")

st.sidebar.markdown("""
<style>
.doc-pill {
background-color:#1f242b;
padding:6px 12px;
border-radius:10px;
margin-bottom:6px;
font-size:12px;
color:#e6edf3;
border:1px solid #30363d;
}
</style>
""", unsafe_allow_html=True)

pdf_list = get_sidebar_doc_entries()


MAX_DOCS_SIDEBAR = 200

if not pdf_list:
    st.sidebar.caption("Nenhum documento na base filtrada atual.")

for i, (doc_base, pdf) in enumerate(sorted(pdf_list, key=lambda item: (item[0], item[1].name))[:MAX_DOCS_SIDEBAR]):

    col_btn, col_name = st.sidebar.columns([1,8])

    with col_btn:

        if st.button(
            "\U0001F4C2",
            key=f"open_{doc_base}_{i}_{pdf.name}"
        ):

            try:

                if sys.platform == "win32":
                    os.startfile(pdf)

                elif sys.platform == "darwin":
                    subprocess.Popen(["open", str(pdf)])

                else:
                    subprocess.Popen(["xdg-open", str(pdf)])

            except Exception as e:

                st.sidebar.error(f"Erro ao abrir arquivo: {e}")

    with col_name:

        label = f"{doc_base} - {pdf.name}" if st.session_state.get("multi_base_enabled") else pdf.name

        st.markdown(
            f"<div class='doc-pill'>{escape(label)}</div>",
            unsafe_allow_html=True
        )


# =====================================================
# LAYOUT CENTRAL
# =====================================================

main_col, config_col = st.columns([3,1])


# =====================================================
# CONFIGURAES RAG
# =====================================================

with config_col:

    st.markdown("## \u2699\ufe0f Parametros do RAG")

    DEFAULT_RAG = {
        "top_k_retrieval": 60,
        "top_k_final": 20,
        "alpha_semantic": 0.7,
        "beta_lexical": 0.3,
        "max_context_chars": 20000,
        "adaptive_tuning_enabled": True,
        "max_query_variants": 4,
        "hybrid_profiles_per_query": 1,
        "max_hybrid_searches": 8,
    }

    for key,value in DEFAULT_RAG.items():

        if key not in st.session_state:
            st.session_state[key] = value

    st.session_state.pop("fallback_on_hybrid_slow", None)

    if st.button("\U0001F504 Restaurar Default (RAG)"):

        for key,value in DEFAULT_RAG.items():
            st.session_state[key] = value

    if st.button("\U0001F9F9 Limpar Chat"):

        st.session_state.chat_history = []
        st.session_state.evidence_state = {}

    st.checkbox("Autoajuste adaptativo por pergunta", key="adaptive_tuning_enabled")
    adaptive_locked = bool(st.session_state.get("adaptive_tuning_enabled", True))

    if adaptive_locked:
        st.caption(
            "Autoajuste ativo: parametros manuais de busca ficam inoperantes."
        )

    st.slider(
        "Top K Retrieval",
        10,
        150,
        key="top_k_retrieval",
        disabled=adaptive_locked,
    )
    st.slider(
        "Top K Final",
        5,
        60,
        key="top_k_final",
        disabled=adaptive_locked,
    )
    st.slider(
        "Peso Semantico (alpha)",
        0.0,
        1.0,
        key="alpha_semantic",
        disabled=adaptive_locked,
    )
    st.slider(
        "Peso Lexical (beta)",
        0.0,
        1.0,
        key="beta_lexical",
        disabled=adaptive_locked,
    )
    st.slider(
        "Max Contexto",
        5000,
        50000,
        step=1000,
        key="max_context_chars",
        disabled=adaptive_locked,
    )
    st.slider(
        "Variacoes de consulta (Hybrid)",
        1,
        20,
        key="max_query_variants",
        disabled=adaptive_locked,
    )
    st.slider(
        "Perfis hibridos por consulta",
        1,
        3,
        key="hybrid_profiles_per_query",
        disabled=adaptive_locked,
    )
    st.slider(
        "Limite total de buscas hibridas",
        1,
        60,
        key="max_hybrid_searches",
        disabled=adaptive_locked,
    )

    st.markdown("---")
    st.markdown("## Memoria")

    memory_feedback = st.session_state.pop("memory_clear_feedback", None)
    if memory_feedback:
        st.success(memory_feedback)

    st.checkbox(
        "Usar memoria nas respostas",
        key="memory_enabled",
        help="Inclui memoria curta, resumos persistentes e recuperacao por similaridade no contexto do RAG.",
    )
    memory_disabled = not bool(st.session_state.get("memory_enabled"))
    st.checkbox(
        "Memoria curta da conversa atual",
        key="memory_short_enabled",
        disabled=memory_disabled,
    )
    st.checkbox(
        "Memoria persistente recuperavel",
        key="memory_persistent_enabled",
        disabled=memory_disabled,
    )
    st.checkbox(
        "Gerar resumo persistente apos cada resposta",
        key="memory_auto_summarize",
        disabled=memory_disabled,
    )
    st.slider(
        "Turnos recentes",
        1,
        10,
        key="memory_short_turns",
        disabled=memory_disabled or not st.session_state.get("memory_short_enabled"),
    )
    st.slider(
        "Memorias recuperadas",
        1,
        12,
        key="memory_retrieval_limit",
        disabled=memory_disabled or not st.session_state.get("memory_persistent_enabled"),
    )
    try:
        stats = memory_stats()
        st.caption(
            f"Memoria salva: {stats['turns']} conversa(s), "
            f"{stats['summaries']} resumo(s)."
        )
    except Exception as exc:
        st.caption(f"Memoria indisponivel: {exc}")

    if st.button("Limpar memoria persistente"):
        removed = clear_memory()
        st.session_state.memory_clear_feedback = (
            f"Memoria limpa ({len(removed)} arquivo(s) removido(s))."
        )
        st.rerun()

    st.markdown("---")
    st.markdown("## \U0001F9E0 Prompt do Sistema")

    prompt_save_feedback = st.session_state.pop("prompt_save_feedback", None)
    if prompt_save_feedback:
        st.success(prompt_save_feedback)

    if "prompt_custom" not in st.session_state:
        st.session_state.prompt_custom = DEFAULT_PROMPT_TEMPLATE

    prompt_template_pending = st.session_state.pop(
        "prompt_template_choice_pending",
        None,
    )
    if prompt_template_pending:
        st.session_state.prompt_template_choice = prompt_template_pending

    if st.button("\U0001F504 Restaurar Default (Prompt)"):

        st.session_state.prompt_template_choice = "Juridico"
        st.session_state.prompt_custom = DEFAULT_PROMPT_TEMPLATE

    st.text_area(
        "Editar Prompt:",
        key="prompt_custom",
        height=300
    )

    prompt_template_options = get_prompt_template_options()
    current_prompt_choice = st.session_state.get("prompt_template_choice", "Juridico")
    if current_prompt_choice not in prompt_template_options:
        st.session_state.prompt_template_choice = "Juridico"

    st.radio(
        "Opcoes de prompt:",
        list(prompt_template_options.keys()),
        key="prompt_template_choice",
        horizontal=True,
        on_change=apply_prompt_template_selection,
        help="Selecione uma opcao para substituir o texto do prompt acima.",
    )

    st.text_input(
        "Salvar prompt como:",
        key="prompt_save_name",
        placeholder="Ex: Parecer tecnico, Auditoria CGU, Resumo executivo",
    )

    if st.button("\U0001F4BE Salvar prompt"):
        try:
            saved_name, saved_path = save_prompt_template(
                st.session_state.prompt_save_name,
                st.session_state.prompt_custom,
            )
            st.session_state.prompt_template_choice_pending = saved_name
            st.session_state.prompt_save_feedback = (
                f"Prompt '{saved_name}' salvo em: {saved_path}"
            )
            st.rerun()
        except ValueError as exc:
            st.warning(str(exc))
        except Exception as exc:
            st.error(f"Erro ao salvar prompt: {exc}")


# =====================================================
# CHAT
# =====================================================

with main_col:

    if "chat_history" not in st.session_state:
        st.session_state.chat_history = []

    if "evidence_state" not in st.session_state:
        st.session_state.evidence_state = {}

    for i,item in enumerate(st.session_state.chat_history):

        st.markdown("### \U0001F464 Pergunta")
        st.markdown(item["question"])

        st.markdown("### \U0001F916 Resposta")
        st.markdown(item["answer"])

        st.markdown(
            f"**Base:** `{item['base']}`  \n"
            f"**Modo:** `{item['modo']}`  \n"
            f"**Data/Hora:** `{item['timestamp']}`"
        )

        timing_history = item.get("timings") or []
        if timing_history:
            st.markdown("##### \u23F1 Tempos da consulta")
            for line in timing_history:
                clean_line = str(line).replace("\u23F1 ", "", 1)
                st.markdown(f":blue[{clean_line}]")

        evidence_text = str(item.get("evidence", "") or "").strip()

        if evidence_text:
            if i not in st.session_state.evidence_state:
                st.session_state.evidence_state[i] = False

            if st.button("\U0001F4CE Mostrar / Ocultar Evidencias", key=f"evidence_{i}"):
                st.session_state.evidence_state[i] = not st.session_state.evidence_state[i]

            if st.session_state.evidence_state[i]:
                st.markdown("#### \U0001F4CE Evidencias")
                st.markdown(evidence_text)

        st.markdown("---")

    query = st.text_area("Pergunte algo",height=80)

    if st.button("Enviar") and query.strip():
        query_for_rag, docs_directive = parse_docs_directive(query)
        if not query_for_rag.strip():
            query_for_rag = query

        meta_answer = answer_conversation_meta_query(
            query_for_rag,
            st.session_state.get("chat_history", []),
        )
        if meta_answer:
            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            timing_lines = [
                "[INFO] Pergunta respondida pela memoria conversacional.",
            ]
            memory_meta = {
                "enabled": bool(st.session_state.get("memory_enabled")),
                "short_turns": min(
                    len(st.session_state.get("chat_history", [])),
                    int(st.session_state.get("memory_short_turns", 4)),
                ),
                "retrieved": 0,
                "meta_query": True,
            }
            st.session_state.chat_history.append({
                "question": query,
                "answer": meta_answer,
                "evidence": "",
                "cited_docs": [],
                "timings": timing_lines,
                "base": "Conversa",
                "modo": "Memoria",
                "timestamp": timestamp,
                "memory": memory_meta,
            })
            st.rerun()

        allowed_docs_scope, docs_directive_resolved, docs_directive_missing = (
            build_docs_directive_scope(docs_directive)
        )

        if st.session_state.get("adaptive_tuning_enabled", True):
            config_override = {"adaptive_tuning_enabled": True}
        else:
            config_override = {
                key: st.session_state[key]
                for key in DEFAULT_RAG.keys()
            }

        timing_lines = []
        status_box = None
        timing_box = None

        if hasattr(st, "status"):
            status_box = st.status(
                "\U0001F6E2 Processando consulta tecnica...",
                state="running",
                expanded=True
            )
        else:
            timing_box = st.empty()

        def on_timing(message):
            timing_lines.append(message)
            clean_line = str(message).replace("\u23F1 ", "", 1)

            if status_box is not None:
                status_box.write(f":blue[{clean_line}]")
            elif timing_box is not None:
                timing_markdown = "\n\n".join(
                    f":blue[{str(line).replace('\u23F1 ', '', 1)}]"
                    for line in timing_lines
                )
                timing_box.markdown(timing_markdown)

        if docs_directive:
            on_timing(format_docs_directive_timing_notice(
                docs_directive_resolved,
                docs_directive_missing,
            ))

        search_route = choose_search_route(
            query_for_rag,
            llm_model=final_llm_model,
            progress_callback=on_timing,
            sqlite_db_path=selected_sqlite_path,
            sqlite_base_name=base,
        )
        route_name = search_route.get("route") or "rag_only"
        routed_rag_query = search_route.get("rag_query") or query_for_rag
        if modo_consulta == "Somente SQLite da base":
            route_name = "sqlite_only"
            routed_rag_query = query_for_rag
            search_route = {
                "route": route_name,
                "reason": "Modo de consulta forçado para SQLite da base.",
                "rag_query": query_for_rag,
                "source": "mode",
            }
        original_query_for_log = query_for_rag
        if search_route.get("source") == "command":
            query_for_rag = routed_rag_query or query_for_rag
        if search_route.get("source") != "command" and routed_rag_query != original_query_for_log:
            on_timing(f"[QUERY][ROUTER] Consulta RAG sugerida: {_short_log_text(routed_rag_query)}")

        memory_context = ""
        memory_meta = {"enabled": False, "short_turns": 0, "retrieved": 0}
        if st.session_state.get("memory_enabled"):
            try:
                memory_context, memory_meta = build_memory_context(
                    routed_rag_query or query_for_rag,
                    chat_history=st.session_state.get("chat_history", []),
                    short_turns=st.session_state.get("memory_short_turns", 4),
                    retrieval_limit=st.session_state.get("memory_retrieval_limit", 5),
                    base=base,
                    include_short=st.session_state.get("memory_short_enabled", True),
                    include_persistent=st.session_state.get("memory_persistent_enabled", True),
                )
                memory_meta["enabled"] = True
                on_timing(
                    "[INFO] Memoria: "
                    f"{memory_meta.get('short_turns', 0)} turno(s) recentes; "
                    f"{memory_meta.get('retrieved', 0)} memoria(s) persistente(s)."
                )
            except Exception as exc:
                memory_context = ""
                memory_meta = {"enabled": False, "error": str(exc)}
                on_timing(f"[WARN] Memoria indisponivel: {exc}")

        try:
            if route_name == "sqlite_only":
                if status_box is not None:
                    status_box.write(
                        ":blue[[INFO] Pergunta roteada para SQLite generico da base.]"
                    )
                on_timing("[INFO] Consulta SQLite da base")
                on_timing(f"[QUERY][SQLITE] {_short_log_text(query_for_rag)}")
                resposta_completa, reranked, routing = answer_universal_sqlite_query(
                    query_for_rag,
                    db_path=selected_sqlite_path,
                    llm_model=final_llm_model,
                    progress_callback=on_timing,
                    label=base,
                    memory_context=memory_context,
                )
                routing["route_source"] = search_route.get("source")
                routing["route_reason"] = search_route.get("reason")
            elif route_name == "cnpj_rag_hybrid":
                if status_box is not None:
                    status_box.write(
                        ":blue[[INFO] Pergunta hibrida: CNPJ SQLite + bases RAG.]"
                    )
                on_timing("[INFO] Consulta CNPJ SQLite")
                cnpj_output, cnpj_reranked, cnpj_routing = answer_cnpj_query(
                    query_for_rag,
                    llm_model=final_llm_model,
                    progress_callback=on_timing,
                )
                cnpj_answer_text, _ = split_answer_and_evidence(cnpj_output)
                hybrid_rag_query = build_hybrid_rag_query(routed_rag_query, cnpj_answer_text)
                on_timing("[INFO] Consulta RAG enriquecida com dados CNPJ")
                on_timing(f"[QUERY][RAG] {_short_log_text(hybrid_rag_query)}")

                if st.session_state.get("multi_base_enabled"):
                    bases_para_consulta = list(selected_bases or [])

                    if status_box is not None:
                        status_box.write(
                            f":blue[[INFO] Modo multibase hibrido: {len(bases_para_consulta)} bases selecionadas.]"
                        )

                    rag_output, rag_reranked, rag_routing = ask_multi_base(
                        query=hybrid_rag_query,
                        base_names=bases_para_consulta,
                        temperature=temperature,
                        llm_model=final_llm_model,
                        config_override=config_override,
                        custom_prompt=st.session_state.prompt_custom,
                        allowed_docs_by_base=allowed_docs_scope,
                        progress_callback=on_timing,
                        memory_context=memory_context,
                    )
                else:
                    if status_box is not None:
                        rag_output, rag_reranked, rag_routing = ask(
                            query=hybrid_rag_query,
                            base_name=base,
                            temperature=temperature,
                            llm_model=final_llm_model,
                            forced_doc=doc_forcado,
                            config_override=config_override,
                            custom_prompt=st.session_state.prompt_custom,
                            allowed_docs=allowed_docs_scope,
                            progress_callback=on_timing,
                            memory_context=memory_context,
                        )
                    else:
                        with st.spinner("\U0001F6E2 Processando consulta tecnica..."):
                            rag_output, rag_reranked, rag_routing = ask(
                                query=hybrid_rag_query,
                                base_name=base,
                                temperature=temperature,
                                llm_model=final_llm_model,
                                forced_doc=doc_forcado,
                                config_override=config_override,
                                custom_prompt=st.session_state.prompt_custom,
                                allowed_docs=allowed_docs_scope,
                                progress_callback=on_timing,
                                memory_context=memory_context,
                            )

                resposta_completa = join_hybrid_outputs(cnpj_output, rag_output)
                reranked = rag_reranked
                routing = {
                    "strategy": "cnpj_rag_hybrid",
                    "cnpj": cnpj_routing,
                    "rag": rag_routing,
                    "cnpj_terms": extract_cnpj_terms_for_rag(cnpj_answer_text),
                    "route_source": search_route.get("source"),
                    "route_reason": search_route.get("reason"),
                }
            elif route_name == "rag_cnpj_hybrid":
                if status_box is not None:
                    status_box.write(
                        ":blue[[INFO] Pergunta hibrida: bases RAG + CNPJ SQLite.]"
                    )
                on_timing("[INFO] Consulta RAG inicial")
                on_timing(f"[QUERY][RAG] {_short_log_text(routed_rag_query)}")

                if st.session_state.get("multi_base_enabled"):
                    bases_para_consulta = list(selected_bases or [])
                    rag_output, rag_reranked, rag_routing = ask_multi_base(
                        query=routed_rag_query,
                        base_names=bases_para_consulta,
                        temperature=temperature,
                        llm_model=final_llm_model,
                        config_override=config_override,
                        custom_prompt=st.session_state.prompt_custom,
                        allowed_docs_by_base=allowed_docs_scope,
                        progress_callback=on_timing,
                        memory_context=memory_context,
                    )
                else:
                    rag_output, rag_reranked, rag_routing = ask(
                        query=routed_rag_query,
                        base_name=base,
                        temperature=temperature,
                        llm_model=final_llm_model,
                        forced_doc=doc_forcado,
                        config_override=config_override,
                        custom_prompt=st.session_state.prompt_custom,
                        allowed_docs=allowed_docs_scope,
                        progress_callback=on_timing,
                        memory_context=memory_context,
                    )

                rag_answer_text, _ = split_answer_and_evidence(rag_output)
                cnpj_query = build_cnpj_query_from_rag(query_for_rag, rag_answer_text)
                on_timing("[INFO] Consulta CNPJ/QSA com achados do RAG")
                on_timing(f"[QUERY][CNPJ] {_short_log_text(cnpj_query)}")
                cnpj_output, cnpj_reranked, cnpj_routing = answer_cnpj_query(
                    cnpj_query,
                    llm_model=final_llm_model,
                    progress_callback=on_timing,
                )
                resposta_completa = join_hybrid_outputs(cnpj_output, rag_output)
                reranked = rag_reranked
                routing = {
                    "strategy": "rag_cnpj_hybrid",
                    "cnpj": cnpj_routing,
                    "rag": rag_routing,
                    "route_source": search_route.get("source"),
                    "route_reason": search_route.get("reason"),
                }
            elif route_name == "cnpj_only":
                if status_box is not None:
                    status_box.write(
                        ":blue[[INFO] Pergunta roteada para base CNPJ SQLite.]"
                    )
                on_timing("[INFO] Consulta CNPJ SQLite")
                on_timing(f"[QUERY][CNPJ] {_short_log_text(query_for_rag)}")
                resposta_completa, reranked, routing = answer_cnpj_query(
                    query_for_rag,
                    llm_model=final_llm_model,
                    progress_callback=on_timing,
                )
                routing["route_source"] = search_route.get("source")
                routing["route_reason"] = search_route.get("reason")
            elif route_name == "anm_only":
                if status_box is not None:
                    status_box.write(
                        ":blue[[INFO] Pergunta roteada para base ANM SQLite.]"
                    )
                on_timing("[INFO] Consulta ANM SQLite")
                on_timing(f"[QUERY][ANM] {_short_log_text(query_for_rag)}")
                resposta_completa, reranked, routing = answer_anm_query(
                    query_for_rag,
                    llm_model=final_llm_model,
                    progress_callback=on_timing,
                )
                routing["route_source"] = search_route.get("source")
                routing["route_reason"] = search_route.get("reason")
            elif st.session_state.get("multi_base_enabled"):
                bases_para_consulta = list(selected_bases or [])

                if status_box is not None:
                    status_box.write(
                        f":blue[[INFO] Modo multibase unificado: {len(bases_para_consulta)} bases selecionadas.]"
                    )

                on_timing(f"[QUERY][RAG] {_short_log_text(routed_rag_query)}")
                resposta_completa, reranked, routing = ask_multi_base(
                    query=routed_rag_query,
                    base_names=bases_para_consulta,
                    temperature=temperature,
                    llm_model=final_llm_model,
                    config_override=config_override,
                    custom_prompt=st.session_state.prompt_custom,
                    allowed_docs_by_base=allowed_docs_scope,
                    progress_callback=on_timing,
                    memory_context=memory_context,
                )
                routing["route_source"] = search_route.get("source")
                routing["route_reason"] = search_route.get("reason")
            else:
                on_timing(f"[QUERY][RAG] {_short_log_text(routed_rag_query)}")
                if status_box is not None:
                    resposta_completa,reranked,routing = ask(
                        query=routed_rag_query,
                        base_name=base,
                        temperature=temperature,
                        llm_model=final_llm_model,
                        forced_doc=doc_forcado,
                        config_override=config_override,
                        custom_prompt=st.session_state.prompt_custom,
                        allowed_docs=allowed_docs_scope,
                        progress_callback=on_timing,
                        memory_context=memory_context,
                    )
                else:
                    with st.spinner("\U0001F6E2 Processando consulta tecnica..."):
                        resposta_completa,reranked,routing = ask(
                            query=routed_rag_query,
                            base_name=base,
                            temperature=temperature,
                            llm_model=final_llm_model,
                            forced_doc=doc_forcado,
                            config_override=config_override,
                            custom_prompt=st.session_state.prompt_custom,
                            allowed_docs=allowed_docs_scope,
                            progress_callback=on_timing,
                            memory_context=memory_context,
                        )
                routing["route_source"] = search_route.get("source")
                routing["route_reason"] = search_route.get("reason")

            if status_box is not None:
                status_box.update(
                    label="\u2705 Consulta concluida.",
                    state="complete",
                    expanded=False
                )
        except Exception:
            if status_box is not None:
                status_box.update(
                    label="\u274C Erro durante a consulta.",
                    state="error",
                    expanded=True
                )
            raise

        resposta_texto, evidencia_texto = split_answer_and_evidence(resposta_completa)
        if docs_directive:
            docs_notice = format_docs_directive_notice(
                docs_directive_resolved,
                docs_directive_missing,
            )
            resposta_texto = f"{docs_notice}\n\n{resposta_texto}"

        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        if routing.get("strategy") == "cnpj_rag_hybrid":
            modo_final = "CNPJ + RAG"
        elif routing.get("strategy") == "rag_cnpj_hybrid":
            modo_final = "RAG + CNPJ"
        elif routing.get("strategy") == "cnpj_relationship_graph":
            modo_final = "CNPJ Grafo societario"
        elif routing.get("strategy") == "cnpj_person_qsa":
            modo_final = "CNPJ QSA"
        elif routing.get("strategy") == "cnpj_company_profile":
            modo_final = "CNPJ Perfil cadastral"
        elif routing.get("strategy") in {"cnpj_sqlite", "cnpj_sqlite_llm_intent"}:
            modo_final = "CNPJ SQLite"
        elif routing.get("strategy") == "anm_sqlite":
            modo_final = "ANM SQLite"
        elif routing.get("strategy") == "sqlite_universal":
            modo_final = "SQLite da base"
        elif st.session_state.get("multi_base_enabled"):
            modo_final = "Multibase"
        elif modo_consulta == "Forcar Documento Especifico":
            modo_final = f"Forcado: {doc_forcado}"
        else:
            modo_final = modo_consulta

        st.session_state.chat_history.append({
            "question": query,
            "answer": resposta_texto,
            "evidence": evidencia_texto,
            "cited_docs": build_cited_docs_from_results(reranked),
            "timings": timing_lines,
            "base": ", ".join(selected_bases or []) if st.session_state.get("multi_base_enabled") else base,
            "modo": modo_final,
            "timestamp": timestamp,
            "memory": memory_meta,
        })

        if st.session_state.get("memory_enabled"):
            try:
                memory_base = (
                    ", ".join(selected_bases or [])
                    if st.session_state.get("multi_base_enabled")
                    else base
                )
                append_conversation_turn(
                    query,
                    resposta_texto,
                    base=memory_base,
                    mode=modo_final,
                    metadata={
                        "route": route_name,
                        "memory": memory_meta,
                    },
                )
                if st.session_state.get("memory_auto_summarize"):
                    append_memory_summary(
                        query,
                        resposta_texto,
                        base=memory_base,
                        mode=modo_final,
                        llm_model=final_llm_model,
                    )
            except Exception as exc:
                st.warning(f"Nao foi possivel salvar memoria: {exc}")

        st.rerun()


